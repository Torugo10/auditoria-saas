# 
# SISTEMA CONTÁBIL AUTOMATIZADO: Auditoria Fiscal, Itens e Controle Financeiro
# Versão: 8.3.23 (Sistema de Suporte Integrado + Totalizadores + Financeiro)
# Desenvolvido para: Victor Hugo - Contador/Empresário Simples Nacional
# Data: 19/03/2026
# 

import os
import streamlit as st
import pandas as pd
import numpy as np
import sqlite3
import hashlib
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import plotly.express as px
import plotly.graph_objects as go

# Database path configuration for Railway
DB_PATH = os.getenv('DATABASE_PATH', '/tmp/auditoria_multi_tenant.db')
os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)

# 
# 🔐 CONFIGURAÇÃO INICIAL
# 

st.set_page_config(page_title="Auditoria Contábil Pro", layout="wide", page_icon="📊")

# Early render — confirms the app reached Streamlit's execution loop.
# Visible in the browser and proves the process is alive before DB init.
print("✅ [STARTUP] st.set_page_config() concluído — Streamlit está respondendo.")

if 'autenticado' not in st.session_state:
    st.session_state.autenticado = False
    st.session_state.usuario = None
    st.session_state.tipo_usuario = None
    st.session_state.cnpj_cpf = None
    st.session_state.perfil = None
    st.session_state.usuario_id = None

# 
# 🗄️ INICIALIZAÇÃO DO BANCO DE DADOS
# 

def inicializar_banco_dados():
    """Cria o banco de dados e todas as tabelas necessárias caso não existam.
    Deve ser chamada antes de qualquer outra operação de banco de dados."""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()

        c.execute('''
        CREATE TABLE IF NOT EXISTS administradores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            login TEXT UNIQUE NOT NULL,
            senha_hash TEXT NOT NULL,
            email TEXT,
            nome_completo TEXT,
            ativo INTEGER DEFAULT 1,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        ''')

        c.execute('''
        CREATE TABLE IF NOT EXISTS gerentes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            admin_id INTEGER NOT NULL,
            login TEXT UNIQUE NOT NULL,
            senha_hash TEXT NOT NULL,
            email TEXT,
            cnpj_cpf TEXT,
            nome_completo TEXT,
            ativo INTEGER DEFAULT 1,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(admin_id) REFERENCES administradores(id)
        )
        ''')

        c.execute('''
        CREATE TABLE IF NOT EXISTS usuarios_finais (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            gerente_id INTEGER NOT NULL,
            admin_id INTEGER,
            login TEXT UNIQUE NOT NULL,
            senha_hash TEXT NOT NULL,
            email TEXT,
            cnpj_cpf TEXT,
            nome_completo TEXT,
            perfil TEXT DEFAULT 'usuario_final',
            ativo INTEGER DEFAULT 1,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(gerente_id) REFERENCES gerentes(id)
        )
        ''')

        c.execute('''
        CREATE TABLE IF NOT EXISTS contas_receber (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario_id INTEGER,
            cnpj_cpf TEXT,
            login TEXT,
            valor_devido REAL,
            data_vencimento DATE,
            status TEXT DEFAULT 'ABERTO',
            cfop TEXT,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            recorrente_id INTEGER
        )
        ''')

        c.execute('''
        CREATE TABLE IF NOT EXISTS contas_recorrentes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            admin_id INTEGER,
            gerente_id INTEGER,
            cnpj_cpf_gerente TEXT,
            usuario_id INTEGER,
            login_usuario TEXT,
            cnpj_cpf_usuario TEXT,
            valor_fixo REAL NOT NULL,
            dia_vencimento INTEGER NOT NULL,
            descricao TEXT,
            ativa INTEGER DEFAULT 1,
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        ''')

        c.execute('''
        CREATE TABLE IF NOT EXISTS logs_auditoria (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario_id INTEGER,
            tipo_usuario TEXT,
            cnpj_cpf TEXT,
            acao TEXT NOT NULL,
            dados_novos TEXT,
            dados_antigos TEXT,
            data_hora TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        ''')

        c.execute('''
        CREATE TABLE IF NOT EXISTS aliquotas_padrao (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cfop TEXT NOT NULL,
            descricao TEXT,
            aliquota_icms_esperada REAL,
            aliquota_pis_esperada REAL,
            aliquota_cofins_esperada REAL,
            aliquota_ipi_esperada REAL,
            regime TEXT DEFAULT 'Simples Nacional',
            anexo TEXT DEFAULT 'III',
            criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
        ''')

        c.execute('''
        CREATE TABLE IF NOT EXISTS tickets_suporte (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario_id INTEGER NOT NULL,
            usuario_login TEXT NOT NULL,
            usuario_email TEXT,
            tipo_problema TEXT NOT NULL,
            assunto TEXT NOT NULL,
            descricao TEXT NOT NULL,
            prioridade TEXT DEFAULT 'MEDIA',
            status TEXT DEFAULT 'ABERTO',
            data_criacao TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            data_resposta TIMESTAMP,
            resposta_admin TEXT,
            admin_id INTEGER
        )
        ''')

        conn.commit()
        conn.close()
        return True, None
    except Exception as e:
        return False, str(e)

# Inicializar banco de dados antes de qualquer outra operação
print("🚀 [STARTUP] Iniciando aplicação Auditoria Contábil Pro...")
try:
    _db_ok, _db_erro = inicializar_banco_dados()
    if not _db_ok:
        print(f"❌ [STARTUP] Falha na inicialização do banco de dados: {_db_erro}")
        st.error(f"❌ Não foi possível inicializar o banco de dados: {_db_erro}")
        st.info("Por favor, verifique as permissões de escrita no diretório da aplicação e tente novamente.")
        st.stop()
    else:
        print("✅ [STARTUP] Banco de dados inicializado com sucesso.")
except Exception as _db_init_error:
    print(f"❌ [STARTUP] Erro crítico ao inicializar o banco de dados: {str(_db_init_error)}")
    st.error(f"❌ Erro crítico ao inicializar o banco de dados: {str(_db_init_error)}")
    st.info("Por favor, verifique as permissões de escrita no diretório da aplicação e tente novamente.")
    st.stop()

# 
# 🔐 FUNÇÕES DE AUTENTICAÇÃO
# 

def autenticar_usuario(login, senha):
    """Autenticação com validação de bloqueio"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    
    senha_hash = hashlib.sha256(senha.encode()).hexdigest()
    
    # Admin
    c.execute('SELECT id, ativo FROM administradores WHERE login = ? AND senha_hash = ?', 
             (login, senha_hash))
    admin = c.fetchone()
    if admin:
        usuario_id, ativo = admin
        if not ativo:
            conn.close()
            registrar_log(0, 'desconhecido', '', 'LOGIN_BLOQUEADO', f'Tentativa: {login}')
            return None
        conn.close()
        return {'usuario_id': usuario_id, 'tipo': 'admin', 'login': login, 'cnpj_cpf': None}
    
    # Gerente
    c.execute('SELECT id, cnpj_cpf, ativo FROM gerentes WHERE login = ? AND senha_hash = ?', 
             (login, senha_hash))
    gerente = c.fetchone()
    if gerente:
        usuario_id, cnpj_cpf, ativo = gerente
        if not ativo:
            conn.close()
            registrar_log(0, 'desconhecido', '', 'LOGIN_BLOQUEADO', f'Tentativa: {login}')
            return None
        conn.close()
        return {'usuario_id': usuario_id, 'tipo': 'gerente', 'login': login, 'cnpj_cpf': cnpj_cpf}
    
    # Usuário Final
    c.execute('SELECT id, cnpj_cpf, perfil, ativo FROM usuarios_finais WHERE login = ? AND senha_hash = ?', 
             (login, senha_hash))
    usuario = c.fetchone()
    if usuario:
        usuario_id, cnpj_cpf, perfil, ativo = usuario
        if not ativo:
            conn.close()
            registrar_log(0, 'desconhecido', '', 'LOGIN_BLOQUEADO', f'Tentativa: {login}')
            return None
        conn.close()
        return {'usuario_id': usuario_id, 'tipo': 'usuario_final', 'login': login, 
               'cnpj_cpf': cnpj_cpf, 'perfil': perfil}
    
    conn.close()
    return None

def validar_sessao_ativa_forcada():
    """Validação CRÍTICA: Executa SEMPRE"""
    if not st.session_state.autenticado:
        return False
    
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        if st.session_state.tipo_usuario == 'admin':
            c.execute('SELECT ativo FROM administradores WHERE id = ? AND login = ?', 
                     (st.session_state.usuario_id, st.session_state.usuario))
        elif st.session_state.tipo_usuario == 'gerente':
            c.execute('SELECT ativo FROM gerentes WHERE id = ? AND login = ?', 
                     (st.session_state.usuario_id, st.session_state.usuario))
        else:
            c.execute('SELECT ativo FROM usuarios_finais WHERE id = ? AND login = ?', 
                     (st.session_state.usuario_id, st.session_state.usuario))
        
        resultado = c.fetchone()
        conn.close()
        
        if not resultado or resultado[0] == 0:
            st.session_state.autenticado = False
            st.session_state.usuario = None
            st.session_state.tipo_usuario = None
            st.session_state.cnpj_cpf = None
            st.session_state.usuario_id = None
            
            st.error("❌ ACESSO BLOQUEADO: Sua conta foi inativada.")
            st.stop()
        
        return True
    except Exception as e:
        st.session_state.autenticado = False
        st.stop()

def bloquear_usuarios_atrasados():
    """Bloqueio automático por atraso de pagamento (>5 dias)"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='contas_receber'")
        if not c.fetchone():
            conn.close()
            return
        
        data_limite = (datetime.now() - timedelta(days=5)).date()
        
        c.execute('''
        SELECT DISTINCT usuario_id FROM contas_receber 
        WHERE status = 'ABERTO' AND data_vencimento < ? AND usuario_id IS NOT NULL
        ''', (data_limite,))
        
        usuarios_atrasados = c.fetchall()
        
        for (usuario_id,) in usuarios_atrasados:
            c.execute('SELECT ativo FROM usuarios_finais WHERE id = ?', (usuario_id,))
            resultado = c.fetchone()
            
            if resultado and resultado[0] == 1:
                c.execute('UPDATE usuarios_finais SET ativo = 0 WHERE id = ?', (usuario_id,))
                
                c.execute('SELECT login FROM usuarios_finais WHERE id = ?', (usuario_id,))
                login_result = c.fetchone()
                login = login_result[0] if login_result else 'desconhecido'
                
                registrar_log(0, 'sistema', '', 'BLOQUEIO_AUTOMATICO', 
                             f'Usuário {login} bloqueado por atraso')
        
        conn.commit()
        conn.close()
    except:
        pass

def registrar_log(usuario_id, tipo_usuario, cnpj_cpf, acao, dados_novos=""):
    """Registra ação no log de auditoria"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        c.execute('''
        INSERT INTO logs_auditoria (usuario_id, tipo_usuario, cnpj_cpf, acao, dados_novos, data_hora)
        VALUES (?, ?, ?, ?, ?, ?)
        ''', (usuario_id, tipo_usuario, cnpj_cpf, acao, dados_novos, datetime.now()))
        
        conn.commit()
        conn.close()
    except:
        pass

# 
# 💳 FUNÇÕES PARA CONTAS RECORRENTES (FIXAS MENSAIS)
# 

def gerar_contas_mensais_recorrentes():
    """Gera automaticamente contas a receber do mês baseado em contas recorrentes"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        hoje = datetime.now().date()
        primeiro_dia = hoje.replace(day=1)
        
        if hoje != primeiro_dia:
            return
        
        c.execute('''
        SELECT id, usuario_id, cnpj_cpf_usuario, login_usuario, 
               valor_fixo, dia_vencimento, descricao, gerente_id
        FROM contas_recorrentes 
        WHERE ativa = 1
        ''')
        
        contas_recorrentes = c.fetchall()
        
        for conta in contas_recorrentes:
            (recorrente_id, usuario_id, cnpj_cpf, login, valor, 
             dia_venc, descricao, gerente_id) = conta
            
            try:
                data_venc = hoje.replace(day=dia_venc)
            except ValueError:
                if dia_venc > 28:
                    data_venc = (hoje.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
                else:
                    data_venc = hoje.replace(day=dia_venc)
            
            c.execute('''
            SELECT id FROM contas_receber 
            WHERE usuario_id = ? AND status = 'ABERTO'
            AND strftime('%Y-%m', data_vencimento) = ?
            ''', (usuario_id, data_venc.strftime('%Y-%m')))
            
            if not c.fetchone():
                c.execute('''
                INSERT INTO contas_receber 
                (usuario_id, cnpj_cpf, login, valor_devido, data_vencimento, 
                 status, criado_em, recorrente_id)
                VALUES (?, ?, ?, ?, ?, 'ABERTO', ?, ?)
                ''', (usuario_id, cnpj_cpf, login, valor, data_venc, 
                      datetime.now(), recorrente_id))
                
                registrar_log(0, 'sistema', '', 'CONTA_RECORRENTE_GERADA',
                            f'Conta de R$ {valor:.2f} para {login} - Vence em {data_venc}')
        
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"❌ Erro ao gerar contas recorrentes: {str(e)}")

def criar_conta_recorrente(admin_id, gerente_id, cnpj_cpf_gerente, usuario_id, 
                          login_usuario, cnpj_cpf_usuario, valor_fixo, 
                          dia_vencimento, descricao=""):
    """Cria uma conta recorrente mensal fixa"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        c.execute('''
        SELECT id FROM usuarios_finais 
        WHERE id = ? AND gerente_id = ? AND cnpj_cpf = ?
        ''', (usuario_id, gerente_id, cnpj_cpf_gerente))
        
        if not c.fetchone():
            conn.close()
            return False, "❌ Usuário não vinculado a este gerente"
        
        c.execute('''
        INSERT INTO contas_recorrentes 
        (admin_id, gerente_id, cnpj_cpf_gerente, usuario_id, login_usuario, 
         cnpj_cpf_usuario, valor_fixo, dia_vencimento, descricao, ativa)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
        ''', (admin_id, gerente_id, cnpj_cpf_gerente, usuario_id, login_usuario,
              cnpj_cpf_usuario, valor_fixo, dia_vencimento, descricao))
        
        conn.commit()
        
        registrar_log(admin_id, 'admin', '', 'CONTA_RECORRENTE_CRIADA',
                    f'Conta recorrente de R$ {valor_fixo:.2f} para {login_usuario} - Vence no dia {dia_vencimento}')
        
        conn.close()
        return True, f"✅ Conta recorrente criada! Será gerada todo dia {dia_vencimento} do mês"
    
    except Exception as e:
        return False, f"❌ Erro: {str(e)}"

def listar_contas_recorrentes(gerente_id=None, admin_id=None):
    """Lista contas recorrentes por gerente ou admin"""
    try:
        conn = sqlite3.connect(DB_PATH)
        
        if gerente_id:
            df = pd.read_sql_query(
                '''SELECT id, login_usuario, valor_fixo, dia_vencimento, descricao, ativa
                   FROM contas_recorrentes 
                   WHERE gerente_id = ? AND ativa = 1
                   ORDER BY login_usuario''',
                conn, params=(gerente_id,))
        elif admin_id:
            df = pd.read_sql_query(
                '''SELECT id, login_usuario, valor_fixo, dia_vencimento, 
                   descricao, gerente_id, cnpj_cpf_gerente, ativa
                   FROM contas_recorrentes 
                   WHERE admin_id = ? AND ativa = 1
                   ORDER BY gerente_id, login_usuario''',
                conn, params=(admin_id,))
        else:
            df = pd.DataFrame()
        
        conn.close()
        return df
    
    except Exception as e:
        return pd.DataFrame()

def desativar_conta_recorrente(recorrente_id, admin_id):
    """Desativa uma conta recorrente"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        c.execute('''
        UPDATE contas_recorrentes SET ativa = 0 
        WHERE id = ? AND admin_id = ?
        ''', (recorrente_id, admin_id))
        
        conn.commit()
        conn.close()
        
        registrar_log(admin_id, 'admin', '', 'CONTA_RECORRENTE_DESATIVADA',
                    f'Conta recorrente ID {recorrente_id} desativada')
        
        return True, "✅ Conta recorrente desativada"
    except Exception as e:
        return False, f"❌ Erro: {str(e)}"

# 
# 🎯 FUNÇÕES DE CONFORMIDADE TRIBUTÁRIA
# 

def criar_tabelas_conformidade():
    """Cria tabelas de referência para análise tributária"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='aliquotas_padrao'")
        if not c.fetchone():
            c.execute('''
            CREATE TABLE aliquotas_padrao (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cfop TEXT NOT NULL,
                descricao TEXT,
                aliquota_icms_esperada REAL,
                aliquota_pis_esperada REAL,
                aliquota_cofins_esperada REAL,
                aliquota_ipi_esperada REAL,
                regime TEXT DEFAULT 'Simples Nacional',
                anexo TEXT DEFAULT 'III',
                criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            ''')
            
            dados_aliquotas = [
                ('1500', 'Compra para revenda', 0.0, 0.65, 3.0, 0.0),
                ('1900', 'Compra com ST', 8.43, 0.65, 3.0, 0.0),
                ('2100', 'Devolução de venda', 0.0, 0.65, 3.0, 0.0),
                ('2350', 'Devolução de compra com ST', 0.0, 0.0, 0.0, 0.0),
                ('5102', 'Venda com ST', 4.5, 0.65, 3.0, 0.0),
            ]
            
            for cfop, desc, icms, pis, cofins, ipi in dados_aliquotas:
                c.execute('''
                INSERT INTO aliquotas_padrao 
                (cfop, descricao, aliquota_icms_esperada, aliquota_pis_esperada, aliquota_cofins_esperada, aliquota_ipi_esperada)
                VALUES (?, ?, ?, ?, ?, ?)
                ''', (cfop, desc, icms, pis, cofins, ipi))
            
            conn.commit()
    except Exception as e:
        pass
    finally:
        conn.close()

def calcular_aliquota_efetiva(valor_base, valor_imposto):
    """Calcula alíquota efetiva realizada"""
    if valor_base == 0 or valor_base is None:
        return 0.0
    try:
        return (float(valor_imposto) / float(valor_base)) * 100
    except:
        return 0.0

def analisar_conformidade_operacoes(usuario_id, periodo_inicio, periodo_fim):
    """Analisa conformidade tributária e retorna desvios e recomendações"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        analise_resultados = {
            'operacoes_auditadas': 0,
            'desvios_detectados': 0,
            'economia_potencial': 0.0,
            'alertas': [],
            'recomendacoes': []
        }
        
        c.execute('SELECT cfop, aliquota_icms_esperada FROM aliquotas_padrao')
        aliquotas_ref = {row[0]: row[1] for row in c.fetchall()}
        
        c.execute('''
        SELECT DISTINCT cfop, SUM(CAST(valor_devido AS REAL)) as base_total
        FROM contas_receber
        WHERE usuario_id = ? AND data_vencimento BETWEEN ? AND ? AND status = 'ABERTO'
        GROUP BY cfop
        ''', (usuario_id, periodo_inicio, periodo_fim))
        
        operacoes = c.fetchall()
        
        for cfop, base_total in operacoes:
            if cfop and base_total:
                analise_resultados['operacoes_auditadas'] += 1
                
                aliq_esperada = aliquotas_ref.get(cfop, 0.0)
                
                analise_resultados['recomendacoes'].append({
                    'tipo': 'MONITORAMENTO_CFOP',
                    'cfop': str(cfop),
                    'aliquota_esperada': round(aliq_esperada, 2),
                    'base_valores': round(base_total, 2),
                    'descricao': f'CFOP {cfop} monitorado - Alíquota esperada: {aliq_esperada}%'
                })
        
        conn.close()
        return analise_resultados
    
    except Exception as e:
        return {
            'operacoes_auditadas': 0,
            'desvios_detectados': 0,
            'economia_potencial': 0.0,
            'alertas': [],
            'recomendacoes': []
        }

def gerar_score_conformidade(usuario_id, periodo_inicio, periodo_fim):
    """Gera score de conformidade tributária (0-100)"""
    analise = analisar_conformidade_operacoes(usuario_id, periodo_inicio, periodo_fim)
    
    if analise['operacoes_auditadas'] == 0:
        return 100.0
    
    return 85.0

# 
# 🆘 FUNÇÕES DE SUPORTE INTEGRADO
# 

def criar_tabela_tickets():
    """Cria tabela de tickets de suporte se não existir"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        c.execute('''
        CREATE TABLE IF NOT EXISTS tickets_suporte (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario_id INTEGER NOT NULL,
            usuario_login TEXT NOT NULL,
            usuario_email TEXT,
            tipo_problema TEXT NOT NULL,
            assunto TEXT NOT NULL,
            descricao TEXT NOT NULL,
            prioridade TEXT DEFAULT 'MEDIA',
            status TEXT DEFAULT 'ABERTO',
            data_criacao TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            data_resposta TIMESTAMP,
            resposta_admin TEXT,
            admin_id INTEGER,
            FOREIGN KEY(usuario_id) REFERENCES usuarios_finais(id)
        )
        ''')
        
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"❌ Erro ao criar tabela de tickets: {str(e)}")

def criar_ticket_suporte(usuario_id, usuario_login, usuario_email, tipo_problema, assunto, descricao, prioridade='MEDIA'):
    """Cria um novo ticket de suporte"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        c.execute('''
        INSERT INTO tickets_suporte 
        (usuario_id, usuario_login, usuario_email, tipo_problema, assunto, descricao, prioridade, status)
        VALUES (?, ?, ?, ?, ?, ?, ?, 'ABERTO')
        ''', (usuario_id, usuario_login, usuario_email, tipo_problema, assunto, descricao, prioridade))
        
        conn.commit()
        ticket_id = c.lastrowid
        conn.close()
        
        registrar_log(usuario_id, 'usuario_final', '', 'TICKET_SUPORTE_CRIADO',
                     f'Ticket #{ticket_id} - {assunto}')
        
        return True, f"✅ Ticket #{ticket_id} criado com sucesso! O administrador responderá em breve."
    
    except Exception as e:
        return False, f"❌ Erro ao criar ticket: {str(e)}"

def listar_meus_tickets(usuario_id):
    """Lista todos os tickets do usuário"""
    try:
        conn = sqlite3.connect(DB_PATH)
        df = pd.read_sql_query(
            '''SELECT id, tipo_problema, assunto, status, prioridade, data_criacao, resposta_admin
               FROM tickets_suporte 
               WHERE usuario_id = ? 
               ORDER BY data_criacao DESC''',
            conn, params=(usuario_id,))
        conn.close()
        return df
    except:
        return pd.DataFrame()

def listar_todos_tickets_admin():
    """Lista todos os tickets para o Admin com ordenação por prioridade"""
    try:
        conn = sqlite3.connect(DB_PATH)
        df = pd.read_sql_query(
            '''SELECT id, usuario_id, usuario_login, usuario_email, tipo_problema, assunto, 
                      descricao, status, prioridade, data_criacao, data_resposta, resposta_admin
               FROM tickets_suporte 
               ORDER BY 
                   CASE prioridade 
                       WHEN 'CRITICA' THEN 1 
                       WHEN 'ALTA' THEN 2 
                       WHEN 'MEDIA' THEN 3 
                       WHEN 'BAIXA' THEN 4 
                   END,
                   data_criacao DESC''',
            conn)
        conn.close()
        return df
    except:
        return pd.DataFrame()

def responder_ticket(ticket_id, resposta_admin, admin_id):
    """Responde um ticket (apenas Admin)"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        c.execute('''
        UPDATE tickets_suporte 
        SET status = 'RESPONDIDO', resposta_admin = ?, data_resposta = ?, admin_id = ?
        WHERE id = ?
        ''', (resposta_admin, datetime.now(), admin_id, ticket_id))
        
        conn.commit()
        conn.close()
        
        registrar_log(admin_id, 'admin', '', 'TICKET_RESPONDIDO', f'Ticket #{ticket_id} respondido')
        
        return True, f"✅ Ticket #{ticket_id} respondido!"
    except Exception as e:
        return False, f"❌ Erro ao responder: {str(e)}"

def fechar_ticket(ticket_id, admin_id):
    """Fecha um ticket (apenas Admin)"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        c.execute('''
        UPDATE tickets_suporte 
        SET status = 'FECHADO'
        WHERE id = ?
        ''', (ticket_id,))
        
        conn.commit()
        conn.close()
        
        registrar_log(admin_id, 'admin', '', 'TICKET_FECHADO', f'Ticket #{ticket_id} fechado')
        
        return True, f"✅ Ticket #{ticket_id} fechado!"
    except Exception as e:
        return False, f"❌ Erro: {str(e)}"

def obter_estatisticas_tickets():
    """Obtém estatísticas de tickets"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        c.execute('SELECT COUNT(*) FROM tickets_suporte')
        total = c.fetchone()[0]
        
        c.execute("SELECT COUNT(*) FROM tickets_suporte WHERE status = 'ABERTO'")
        abertos = c.fetchone()[0]
        
        c.execute("SELECT COUNT(*) FROM tickets_suporte WHERE status = 'RESPONDIDO'")
        respondidos = c.fetchone()[0]
        
        c.execute("SELECT COUNT(*) FROM tickets_suporte WHERE status = 'FECHADO'")
        fechados = c.fetchone()[0]
        
        c.execute("SELECT COUNT(*) FROM tickets_suporte WHERE prioridade = 'CRITICA' AND status = 'ABERTO'")
        criticos = c.fetchone()[0]
        
        conn.close()
        
        return {
            'total': total,
            'abertos': abertos,
            'respondidos': respondidos,
            'fechados': fechados,
            'criticos': criticos
        }
    except:
        return {'total': 0, 'abertos': 0, 'respondidos': 0, 'fechados': 0, 'criticos': 0}

# 
# 🔐 FUNÇÕES DE VALIDAÇÃO PARA GERENTES
# 

def validar_usuario_subordinado(gerente_id, usuario_id):
    """Valida se o usuário é subordinado do gerente"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        c.execute('SELECT gerente_id FROM usuarios_finais WHERE id = ?', (int(usuario_id),))
        resultado = c.fetchone()
        conn.close()
        
        if not resultado:
            return False, "❌ Usuário não encontrado"
        
        if int(resultado[0]) != int(gerente_id):
            return False, "❌ Você não tem permissão para gerenciar este usuário"
        
        return True, "Validação OK"
    except Exception as e:
        return False, f"❌ Erro na validação: {str(e)}"

def bloquear_usuario_subordinado(gerente_id, usuario_id):
    """Bloqueia usuário subordinado"""
    valido, mensagem = validar_usuario_subordinado(gerente_id, usuario_id)
    
    if not valido:
        return False, mensagem
    
    try:
        usuario_id = int(usuario_id)
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        c.execute('SELECT login FROM usuarios_finais WHERE id = ?', (usuario_id,))
        login_result = c.fetchone()
        login = login_result[0] if login_result else 'desconhecido'
        
        c.execute('UPDATE usuarios_finais SET ativo = 0 WHERE id = ?', (usuario_id,))
        conn.commit()
        conn.close()
        
        registrar_log(gerente_id, 'gerente', '', 'USUARIO_BLOQUEADO',
                     f'Gerente bloqueou usuário {login} (ID {usuario_id})')
        
        return True, f"✅ Usuário '{login}' bloqueado com sucesso!"
    except Exception as e:
        return False, f"❌ Erro ao bloquear: {str(e)}"

def desbloquear_usuario_subordinado(gerente_id, usuario_id):
    """Desbloqueia usuário subordinado"""
    valido, mensagem = validar_usuario_subordinado(gerente_id, usuario_id)
    
    if not valido:
        return False, mensagem
    
    try:
        usuario_id = int(usuario_id)
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        c.execute('SELECT login FROM usuarios_finais WHERE id = ?', (usuario_id,))
        login_result = c.fetchone()
        login = login_result[0] if login_result else 'desconhecido'
        
        c.execute('UPDATE usuarios_finais SET ativo = 1 WHERE id = ?', (usuario_id,))
        conn.commit()
        conn.close()
        
        registrar_log(gerente_id, 'gerente', '', 'USUARIO_DESBLOQUEADO',
                     f'Gerente desbloqueou usuário {login} (ID {usuario_id})')
        
        return True, f"✅ Usuário '{login}' desbloqueado com sucesso!"
    except Exception as e:
        return False, f"❌ Erro ao desbloquear: {str(e)}"

# 
# 🔐 TELA DE LOGIN
# 

try:
    if not st.session_state.autenticado:
        st.markdown("<h2 style='text-align: center;'>🔐 Acesso Restrito</h2>", unsafe_allow_html=True)
        st.markdown("<p style='text-align: center; color: gray;'>Sistema de Auditoria Fiscal BPO</p>", 
                   unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns([1, 1, 1])
        with col2:
            with st.form("login_form"):
                usuario = st.text_input("Usuário")
                senha = st.text_input("Senha", type="password")
                submit = st.form_submit_button("Entrar", use_container_width=True)
                
                if submit:
                    resultado = autenticar_usuario(usuario, senha)
                    if resultado:
                        st.session_state.autenticado = True
                        st.session_state.usuario = resultado['login']
                        st.session_state.tipo_usuario = resultado['tipo']
                        st.session_state.cnpj_cpf = resultado['cnpj_cpf']
                        st.session_state.perfil = resultado.get('perfil', None)
                        st.session_state.usuario_id = resultado['usuario_id']
                        
                        registrar_log(
                            usuario_id=resultado['usuario_id'],
                            tipo_usuario=resultado['tipo'],
                            cnpj_cpf=resultado['cnpj_cpf'] if resultado['cnpj_cpf'] else '',
                            acao='LOGIN_SUCESSO'
                        )
                        
                        
                    else:
                        st.error("❌ Credenciais inválidas ou conta bloqueada")
                        registrar_log(0, 'desconhecido', '', 'LOGIN_FALHA', f'Usuário: {usuario}')
        
        st.stop()
except st.runtime.scriptrunner.StopException:
    raise
except Exception as _login_err:
    import traceback as _tb
    _login_tb = _tb.format_exc()
    print(f"❌ [STARTUP] Erro na tela de login: {str(_login_err)}")
    print(_login_tb)
    st.error(f"❌ Erro ao renderizar a tela de login: {str(_login_err)}")
    st.code(_login_tb, language="python")
    st.stop()

# ✅ VALIDAÇÃO FORÇADA EXECUTADA SEMPRE
validar_sessao_ativa_forcada()

# ✅ BLOQUEIO AUTOMÁTICO
bloquear_usuarios_atrasados()

# ✅ GERAÇÃO AUTOMÁTICA DE CONTAS RECORRENTES (1º DO MÊS)
gerar_contas_mensais_recorrentes()

# ✅ INICIALIZAR CONFORMIDADE TRIBUTÁRIA
criar_tabelas_conformidade()

# ✅ INICIALIZAR TABELA DE TICKETS
criar_tabela_tickets()

# 
# 📊 FUNÇÕES UTILITÁRIAS
# 

st.markdown("""
    <style>
        [data-testid="stFileUploadDropzone"] div div::before { content: "↕️ Arraste e solte a planilha aqui"; color: #555; font-weight: 600; display: block; margin-bottom: 5px; }
        [data-testid="stFileUploadDropzone"] div div span:first-child { display: none; }
        [data-testid="stFileUploadDropzone"] div div small { display: none; }
    </style>
""", unsafe_allow_html=True)

def load_data(file):
    nome_arquivo = file.name.lower()
    if nome_arquivo.endswith('.csv'):
        try: 
            return pd.read_csv(file, header=None)
        except:
            file.seek(0)
            return pd.read_csv(file, sep=';', encoding='latin1', header=None)
    elif nome_arquivo.endswith(('.xlsx', '.xls')):
        try: 
            return pd.read_excel(file, header=None)
        except: 
            st.error("⚠️ Formato não reconhecido.")
            return None
    return None

def localizar_cabecalho_real(df, chaves_principais, chaves_alternativas=None):
    limite = min(30, len(df))
    for i in range(limite):
        linha = df.iloc[i].astype(str).str.lower().tolist()
        if all(any(palavra.lower() in celula for celula in linha) for palavra in chaves_principais):
            novo_df = df.iloc[i+1:].copy()
            novo_df.columns = df.iloc[i].values
            novo_df.columns = [str(col).strip() for col in novo_df.columns]
            return novo_df
        if chaves_alternativas and all(any(palavra.lower() in celula for celula in linha) for palavra in chaves_alternativas):
            novo_df = df.iloc[i+1:].copy()
            novo_df.columns = df.iloc[i].values
            novo_df.columns = [str(col).strip() for col in novo_df.columns]
            return novo_df
    return df

def parser_cabecalho_dominio(df_raw):
    linha_macro = -1
    linha_detalhe = -1
    for i in range(min(20, len(df_raw))):
        linha_str = df_raw.iloc[i].astype(str).str.lower().tolist()
        if any('icms normal' in val for val in linha_str) and any('pis' in val for val in linha_str):
            linha_macro = i
            linha_detalhe = i + 1
            break
    if linha_macro != -1:
        s_macro = df_raw.iloc[linha_macro].copy().astype(str).replace('nan', np.nan)
        s_detalhe = df_raw.iloc[linha_detalhe].copy().astype(str)
        s_macro = s_macro.ffill()
        novas_colunas = []
        for macro, detalhe in zip(s_macro, s_detalhe):
            m_clean = str(macro).strip().lower()
            d_clean = str(detalhe).strip().lower()
            if any(imp in m_clean for imp in ['icms', 'pis', 'cofins', 'ipi']):
                novas_colunas.append(f"{m_clean}|{d_clean}")
            else: 
                novas_colunas.append(d_clean)
        df_clean = df_raw.iloc[linha_detalhe + 1:].copy()
        df_clean.columns = novas_colunas
        return df_clean.dropna(subset=[col for col in df_clean.columns if 'cfop' in str(col) or 'ncm' in str(col)], thresh=1)
    else:
        df_clean = df_raw.copy()
        for i in range(min(20, len(df_clean))):
            linha = df_clean.iloc[i].astype(str).str.lower().tolist()
            if any(p in celula for celula in linha for p in ['nota', 'produto', 'cfop']):
                df_clean.columns = df_clean.iloc[i].values
                return df_clean.iloc[i+1:].reset_index(drop=True)
        return df_clean

def parser_cabecalho_saida(df_raw):
    linha_macro = -1
    linha_detalhe = -1
    
    for i in range(min(10, len(df_raw))):
        linha_str = df_raw.iloc[i].astype(str).str.lower().tolist()
        if any('icms' in val for val in linha_str) and any('pis' in val for val in linha_str) and any('cofins' in val for val in linha_str):
            linha_macro = i
            linha_detalhe = i + 1
            break
    
    if linha_macro != -1:
        s_macro = df_raw.iloc[linha_macro].copy().astype(str).replace('nan', np.nan)
        s_detalhe = df_raw.iloc[linha_detalhe].copy().astype(str)
        s_macro = s_macro.ffill()
        
        novas_colunas = []
        for macro, detalhe in zip(s_macro, s_detalhe):
            m_clean = str(macro).strip().lower()
            d_clean = str(detalhe).strip().lower()
            
            if any(imp in m_clean for imp in ['icms', 'pis', 'cofins']):
                novas_colunas.append(f"{m_clean}|{d_clean}")
            else:
                novas_colunas.append(d_clean)
        
        df_clean = df_raw.iloc[linha_detalhe + 1:].copy()
        df_clean.columns = novas_colunas
        return df_clean.dropna(subset=[col for col in df_clean.columns if 'cfop' in str(col)], thresh=1)
    else:
        return localizar_cabecalho_real(df_raw, ['cfop', 'nota'], ['cfop', 'produto'])

def converter_valor_br(valor):
    if pd.isna(valor) or valor == '' or str(valor).lower() == 'nan': 
        return 0.0
    if isinstance(valor, (int, float)): 
        return float(valor)
    v_str = str(valor).strip()
    if not any(char.isdigit() for char in v_str): 
        return 0.0
    if ',' in v_str and '.' in v_str: 
        v_str = v_str.replace('.', '').replace(',', '.')
    elif ',' in v_str: 
        v_str = v_str.replace(',', '.')
    try: 
        return float(v_str)
    except: 
        return 0.0

def limpa_nota(valor):
    s = str(valor).strip()
    if s.endswith('.0'): 
        return s[:-2]
    return s

def pegar_coluna_todas_chaves(df, palavras_chave):
    for col in df.columns:
        if all(p.lower() in str(col).lower() for p in palavras_chave):
            dados = df[col]
            if isinstance(dados, pd.DataFrame): 
                return dados.iloc[:, 0]
            return dados
    return pd.Series(0.0, index=df.index, name="ND")

def pegar_coluna_qualquer_chave(df, palavras_chave):
    for col in df.columns:
        if any(p.lower() in str(col).lower() for p in palavras_chave):
            dados = df[col]
            if isinstance(dados, pd.DataFrame): 
                return dados.iloc[:, 0]
            return dados
    return pd.Series(0.0, index=df.index, name="ND")

def formata_br(valor): 
    return f"R$ {valor:,.2f}".replace(',', 'v').replace('.', ',').replace('v', '.')

def formata_pct(valor): 
    return f"{valor:.2f}%"

def exportar_excel_formatado(df, nome_aba="Dados"):
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=nome_aba, index=False)
        
        workbook = writer.book
        worksheet = writer.sheets[nome_aba]
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                if isinstance(cell.value, (int, float)) and cell.value != 0:
                    cell.number_format = '_("R$"* #,##0.00_);_("R$"* (#,##0.00);_("R$"* "-"??_);_(@_)'
        
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        worksheet.freeze_panes = "A2"
    
    output.seek(0)
    return output

# 
# 📦 MÓDULO DE AUDITORIA DE ITENS - ENTRADA
# 

def app_auditoria_itens_entrada():
    st.header("📦 Auditoria de Movimento de Produtos (Itens) - ENTRADA")
    arquivo_itens = st.file_uploader("Upload", type=['csv', 'xlsx', 'xls'], key="up_itens", label_visibility="collapsed")
    
    if arquivo_itens:
        with st.spinner("Compilando árvore hierárquica..."):
            df_raw = load_data(arquivo_itens)
            if df_raw is not None:
                df = parser_cabecalho_dominio(df_raw)
                
                col_nota = pegar_coluna_qualquer_chave(df, ['nota', 'nº'])
                col_prod = pegar_coluna_qualquer_chave(df, ['produto', 'código'])
                col_descricao = pegar_coluna_qualquer_chave(df, ['descrição', 'descricao', 'desc'])
                col_ncm = pegar_coluna_qualquer_chave(df, ['ncm', 'código ncm'])
                col_cfop = pegar_coluna_qualquer_chave(df, ['cfop'])
                col_vlr_prod = pegar_coluna_qualquer_chave(df, ['vlr prod', 'valor do item'])
                col_desconto = pegar_coluna_qualquer_chave(df, ['desconto', 'vlr desc'])
                col_desp = pegar_coluna_qualquer_chave(df, ['desp aces', 'acréscimo'])
                
                if col_cfop.name == "ND":
                    st.error("❌ Chaves primárias não localizadas.")
                    return
                
                df['CFOP_Clean'] = col_cfop.astype(str).str.replace('.0','', regex=False).str.strip()
                df = df[df['CFOP_Clean'].str.contains(r'\d', na=False, regex=True)].copy()
                
                df['Descricao'] = col_descricao.astype(str).str.strip() if col_descricao.name != "ND" else "Não Informado"
                df['NCM'] = col_ncm.astype(str).str.replace('.0','', regex=False).str.strip() if col_ncm.name != "ND" else "Não Informado"
                df['Codigo_Produto'] = col_prod.astype(str).str.strip() if col_prod.name != "ND" else "N/A"
                
                df['Valor_Prod'] = col_vlr_prod.apply(converter_valor_br) if col_vlr_prod.name != "ND" else 0.0
                df['Vlr_Desc'] = col_desconto.apply(converter_valor_br) if col_desconto.name != "ND" else 0.0
                df['Vlr_Desp'] = col_desp.apply(converter_valor_br) if col_desp.name != "ND" else 0.0
                df['Total_Liquido_Produto'] = df['Valor_Prod'] - df['Vlr_Desc'] + df['Vlr_Desp']

                df['CST_ICMS'] = pegar_coluna_todas_chaves(df, ['icms normal', 'cst']).astype(str).str.replace('.0','', regex=False).str.strip()
                df['Base_ICMS'] = pegar_coluna_todas_chaves(df, ['icms normal', 'base']).apply(converter_valor_br)
                df['Alq_ICMS'] = pegar_coluna_todas_chaves(df, ['icms normal', 'alq']).apply(converter_valor_br)
                df['Vlr_ICMS'] = pegar_coluna_todas_chaves(df, ['icms normal', 'valor']).apply(converter_valor_br)

                df['CST_PIS'] = pegar_coluna_todas_chaves(df, ['pis', 'cst']).astype(str).str.replace('.0','', regex=False).str.strip()
                df['Base_PIS'] = pegar_coluna_todas_chaves(df, ['pis', 'base']).apply(converter_valor_br)
                df['Alq_PIS'] = pegar_coluna_todas_chaves(df, ['pis', 'alq']).apply(converter_valor_br)
                df['Vlr_PIS'] = pegar_coluna_todas_chaves(df, ['pis', 'valor']).apply(converter_valor_br)

                df['CST_COFINS'] = pegar_coluna_todas_chaves(df, ['cofins', 'cst']).astype(str).str.replace('.0','', regex=False).str.strip()
                df['Base_COFINS'] = pegar_coluna_todas_chaves(df, ['cofins', 'base']).apply(converter_valor_br)
                df['Alq_COFINS'] = pegar_coluna_todas_chaves(df, ['cofins', 'alq']).apply(converter_valor_br)
                df['Vlr_COFINS'] = pegar_coluna_todas_chaves(df, ['cofins', 'valor']).apply(converter_valor_br)

                st.success("✅ Dados processados!")
                st.divider()

                visao = st.radio("Fluxo Tributário:", ["🟦 ICMS Normal", "🟨 PIS", "🟩 COFINS"], horizontal=True)

                def renderizar_drill_tributario(nome_imposto, sufixo_imposto):
                    col_cst_imposto = f'CST_{sufixo_imposto}'
                    
                    df_resumo = df.groupby(['CFOP_Clean', col_cst_imposto]).agg({'Total_Liquido_Produto': 'sum'}).reset_index()
                    df_resumo.rename(columns={'CFOP_Clean': 'CFOP', col_cst_imposto: 'CST'}, inplace=True)
                    df_resumo['Total_Formatado'] = df_resumo['Total_Liquido_Produto'].apply(formata_br)
                    
                    st.markdown(f"### Consolidado ({nome_imposto})")
                    st.dataframe(df_resumo[['CFOP', 'CST', 'Total_Formatado']], use_container_width=True)
                    
                    col_exp1, col_exp2 = st.columns([4, 1])
                    with col_exp2:
                        excel_consolidado = exportar_excel_formatado(df_resumo[['CFOP', 'CST', 'Total_Liquido_Produto']], f"Consolidado_{sufixo_imposto}")
                        st.download_button(
                            label="📥 Excel",
                            data=excel_consolidado,
                            file_name=f"Consolidado_Entrada_{sufixo_imposto}_{datetime.now().strftime('%d%m%Y')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"download_consolidado_entrada_{sufixo_imposto}"
                        )
                    
                    lista_grupos = df_resumo.apply(lambda row: f"CFOP: {row['CFOP']} | CST: {row['CST']}", axis=1).tolist()
                    if lista_grupos:
                        selecao = st.selectbox(f"Auditar:", lista_grupos, key=f"sel_entrada_{sufixo_imposto}")
                        cfop_sel = selecao.split(" | ")[0].replace("CFOP: ", "")
                        cst_sel = selecao.split(" | ")[1].replace("CST: ", "")
                        
                        df_filtro = df[(df['CFOP_Clean'] == cfop_sel) & (df[col_cst_imposto] == cst_sel)].copy()
                        
                        df_out = pd.DataFrame()
                        df_out["Nota"] = df_filtro[col_nota.name].values if col_nota.name != "ND" else "N/A"
                        df_out["Produto"] = df_filtro[col_prod.name].values if col_prod.name != "ND" else "N/A"
                        df_out["Descrição"] = df_filtro['Descricao'].values
                        df_out["NCM"] = df_filtro['NCM'].values
                        df_out["CST"] = df_filtro[col_cst_imposto].values
                        df_out["Base Calc"] = df_filtro[f'Base_{sufixo_imposto}'].apply(formata_br).values
                        df_out["Alíquota"] = df_filtro[f'Alq_{sufixo_imposto}'].apply(formata_pct).values
                        df_out["Valor Imposto"] = df_filtro[f'Vlr_{sufixo_imposto}'].apply(formata_br).values
                        
                        st.dataframe(df_out, use_container_width=True)
                        
                        # ✅ TOTALIZADOR DINÂMICO - ENTRADA
                        st.divider()
                        st.markdown("### 📊 TOTALIZADOR DO FILTRO APLICADO")
                        
                        col_tot1, col_tot2, col_tot3 = st.columns(3)
                        
                        base_total = df_filtro[f'Base_{sufixo_imposto}'].sum()
                        valor_total = df_filtro[f'Vlr_{sufixo_imposto}'].sum()
                        aliquota_media = (valor_total / base_total * 100) if base_total > 0 else 0
                        
                        with col_tot1:
                            st.metric(
                                f"📊 Base {sufixo_imposto}",
                                formata_br(base_total),
                                delta=f"{len(df_filtro)} operações"
                            )
                        
                        with col_tot2:
                            st.metric(
                                f"💰 Valor {sufixo_imposto}",
                                formata_br(valor_total),
                                delta=f"Alíquota: {aliquota_media:.2f}%"
                            )
                        
                        with col_tot3:
                            st.metric(
                                f"📈 Alíquota Efetiva",
                                f"{aliquota_media:.2f}%",
                                delta=f"Total: {formata_br(base_total + valor_total)}"
                            )
                        
                        st.divider()
                        
                        col_exp3, col_exp4 = st.columns([4, 1])
                        with col_exp4:
                            df_export = pd.DataFrame()
                            df_export["Nota"] = df_filtro[col_nota.name].values if col_nota.name != "ND" else "N/A"
                            df_export["Produto"] = df_filtro[col_prod.name].values if col_prod.name != "ND" else "N/A"
                            df_export["Descrição"] = df_filtro['Descricao'].values
                            df_export["NCM"] = df_filtro['NCM'].values
                            df_export["CST"] = df_filtro[col_cst_imposto].values
                            df_export["Base Calc"] = df_filtro[f'Base_{sufixo_imposto}'].values
                            df_export["Alíquota"] = df_filtro[f'Alq_{sufixo_imposto}'].values
                            df_export["Valor Imposto"] = df_filtro[f'Vlr_{sufixo_imposto}'].values
                            
                            excel_detalhe = exportar_excel_formatado(df_export, f"Detalhe_{sufixo_imposto}")
                            st.download_button(
                                label="📥 Excel",
                                data=excel_detalhe,
                                file_name=f"Detalhe_Entrada_{sufixo_imposto}_{datetime.now().strftime('%d%m%Y')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"download_detalhe_entrada_{sufixo_imposto}"
                            )

                if visao == "🟦 ICMS Normal": renderizar_drill_tributario("ICMS Normal", "ICMS")
                elif visao == "🟨 PIS": renderizar_drill_tributario("PIS", "PIS")
                elif visao == "🟩 COFINS": renderizar_drill_tributario("COFINS", "COFINS")

# 
# 📦 MÓDULO DE AUDITORIA DE ITENS - SAÍDA
# 

def app_auditoria_itens_saida():
    st.header("📦 Auditoria de Movimento de Produtos (Itens) - SAÍDA")
    arquivo_itens = st.file_uploader("Upload", type=['csv', 'xlsx', 'xls'], key="up_itens_saida", label_visibility="collapsed")
    
    if arquivo_itens:
        with st.spinner("Compilando árvore hierárquica..."):
            df_raw = load_data(arquivo_itens)
            if df_raw is not None:
                df = parser_cabecalho_saida(df_raw)
                
                col_nota = pegar_coluna_qualquer_chave(df, ['nota', 'nº'])
                col_prod = pegar_coluna_qualquer_chave(df, ['cód prod', 'código prod', 'código', 'produto'])
                # ✅ REMOVIDO 'desc' - APENAS PARA SAÍDA
                col_descricao = pegar_coluna_qualquer_chave(df, ['descrição prod', 'descrição', 'descricao'])
                col_ncm = pegar_coluna_qualquer_chave(df, ['ncm', 'código ncm'])
                col_cfop = pegar_coluna_qualquer_chave(df, ['cfop'])
                col_vlr_prod = pegar_coluna_qualquer_chave(df, ['valor prod', 'valor do item', 'valor'])
                col_desconto = pegar_coluna_qualquer_chave(df, ['desconto', 'vlr desc'])
                col_desp = pegar_coluna_qualquer_chave(df, ['desp ass', 'desp aces', 'despesas acessórias', 'acréscimo'])
                
                if col_cfop.name == "ND":
                    st.error("❌ Chaves primárias não localizadas.")
                    return
                
                df['CFOP_Clean'] = col_cfop.astype(str).str.replace('.0','', regex=False).str.strip()
                df = df[df['CFOP_Clean'].str.contains(r'\d', na=False, regex=True)].copy()
                
                df['Descricao'] = col_descricao.astype(str).str.strip() if col_descricao.name != "ND" else "Não Informado"
                df['NCM'] = col_ncm.astype(str).str.replace('.0','', regex=False).str.strip() if col_ncm.name != "ND" else "Não Informado"
                df['Codigo_Produto'] = col_prod.astype(str).str.strip() if col_prod.name != "ND" else "N/A"
                
                df['Valor_Prod'] = col_vlr_prod.apply(converter_valor_br) if col_vlr_prod.name != "ND" else 0.0
                df['Vlr_Desc'] = col_desconto.apply(converter_valor_br) if col_desconto.name != "ND" else 0.0
                df['Vlr_Desp'] = col_desp.apply(converter_valor_br) if col_desp.name != "ND" else 0.0
                df['Total_Liquido_Produto'] = df['Valor_Prod'] - df['Vlr_Desc'] + df['Vlr_Desp']

                df['CST_ICMS'] = pegar_coluna_todas_chaves(df, ['icms', 'cst']).astype(str).str.replace('.0','', regex=False).str.strip()
                df['Base_ICMS'] = pegar_coluna_todas_chaves(df, ['icms', 'base']).apply(converter_valor_br)
                df['Alq_ICMS'] = pegar_coluna_todas_chaves(df, ['icms', 'alq']).apply(converter_valor_br)
                df['Vlr_ICMS'] = pegar_coluna_todas_chaves(df, ['icms', 'valor']).apply(converter_valor_br)

                df['CST_PIS'] = pegar_coluna_todas_chaves(df, ['pis', 'cst']).astype(str).str.replace('.0','', regex=False).str.strip()
                df['Base_PIS'] = pegar_coluna_todas_chaves(df, ['pis', 'base']).apply(converter_valor_br)
                df['Alq_PIS'] = pegar_coluna_todas_chaves(df, ['pis', 'alq']).apply(converter_valor_br)
                df['Vlr_PIS'] = pegar_coluna_todas_chaves(df, ['pis', 'valor']).apply(converter_valor_br)

                df['CST_COFINS'] = pegar_coluna_todas_chaves(df, ['cofins', 'cst']).astype(str).str.replace('.0','', regex=False).str.strip()
                df['Base_COFINS'] = pegar_coluna_todas_chaves(df, ['cofins', 'base']).apply(converter_valor_br)
                df['Alq_COFINS'] = pegar_coluna_todas_chaves(df, ['cofins', 'alq']).apply(converter_valor_br)
                df['Vlr_COFINS'] = pegar_coluna_todas_chaves(df, ['cofins', 'valor']).apply(converter_valor_br)

                st.success("✅ Dados processados!")
                st.divider()

                visao = st.radio("Fluxo Tributário:", ["🟦 ICMS Normal", "🟨 PIS", "🟩 COFINS"], horizontal=True, key="visao_saida")

                def renderizar_drill_tributario(nome_imposto, sufixo_imposto):
                    col_cst_imposto = f'CST_{sufixo_imposto}'
                    
                    df_resumo = df.groupby(['CFOP_Clean', col_cst_imposto]).agg({'Total_Liquido_Produto': 'sum'}).reset_index()
                    df_resumo.rename(columns={'CFOP_Clean': 'CFOP', col_cst_imposto: 'CST'}, inplace=True)
                    df_resumo['Total_Formatado'] = df_resumo['Total_Liquido_Produto'].apply(formata_br)
                    
                    st.markdown(f"### Consolidado ({nome_imposto})")
                    st.dataframe(df_resumo[['CFOP', 'CST', 'Total_Formatado']], use_container_width=True)
                    
                    col_exp1, col_exp2 = st.columns([4, 1])
                    with col_exp2:
                        excel_consolidado = exportar_excel_formatado(df_resumo[['CFOP', 'CST', 'Total_Liquido_Produto']], f"Consolidado_{sufixo_imposto}")
                        st.download_button(
                            label="📥 Excel",
                            data=excel_consolidado,
                            file_name=f"Consolidado_Saida_{sufixo_imposto}_{datetime.now().strftime('%d%m%Y')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"download_consolidado_saida_{sufixo_imposto}"
                        )
                    
                    lista_grupos = df_resumo.apply(lambda row: f"CFOP: {row['CFOP']} | CST: {row['CST']}", axis=1).tolist()
                    if lista_grupos:
                        selecao = st.selectbox(f"Auditar:", lista_grupos, key=f"sel_saida_{sufixo_imposto}")
                        cfop_sel = selecao.split(" | ")[0].replace("CFOP: ", "")
                        cst_sel = selecao.split(" | ")[1].replace("CST: ", "")
                        
                        df_filtro = df[(df['CFOP_Clean'] == cfop_sel) & (df[col_cst_imposto] == cst_sel)].copy()
                        
                        df_out = pd.DataFrame()
                        df_out["Nota"] = df_filtro[col_nota.name].values if col_nota.name != "ND" else "N/A"
                        df_out["Produto"] = df_filtro[col_prod.name].values if col_prod.name != "ND" else "N/A"
                        df_out["Descrição"] = df_filtro['Descricao'].values
                        df_out["NCM"] = df_filtro['NCM'].values
                        df_out["CST"] = df_filtro[col_cst_imposto].values
                        df_out["Base Calc"] = df_filtro[f'Base_{sufixo_imposto}'].apply(formata_br).values
                        df_out["Alíquota"] = df_filtro[f'Alq_{sufixo_imposto}'].apply(formata_pct).values
                        df_out["Valor Imposto"] = df_filtro[f'Vlr_{sufixo_imposto}'].apply(formata_br).values
                        
                        st.dataframe(df_out, use_container_width=True)
                        
                        # ✅ TOTALIZADOR DINÂMICO - SAÍDA
                        st.divider()
                        st.markdown("### 📊 TOTALIZADOR DO FILTRO APLICADO")
                        
                        col_tot1, col_tot2, col_tot3 = st.columns(3)
                        
                        base_total = df_filtro[f'Base_{sufixo_imposto}'].sum()
                        valor_total = df_filtro[f'Vlr_{sufixo_imposto}'].sum()
                        aliquota_media = (valor_total / base_total * 100) if base_total > 0 else 0
                        
                        with col_tot1:
                            st.metric(
                                f"📊 Base {sufixo_imposto}",
                                formata_br(base_total),
                                delta=f"{len(df_filtro)} operações"
                            )
                        
                        with col_tot2:
                            st.metric(
                                f"💰 Valor {sufixo_imposto}",
                                formata_br(valor_total),
                                delta=f"Alíquota: {aliquota_media:.2f}%"
                            )
                        
                        with col_tot3:
                            st.metric(
                                f"📈 Alíquota Efetiva",
                                f"{aliquota_media:.2f}%",
                                delta=f"Total: {formata_br(base_total + valor_total)}"
                            )
                        
                        st.divider()
                        
                        col_exp3, col_exp4 = st.columns([4, 1])
                        with col_exp4:
                            df_export = pd.DataFrame()
                            df_export["Nota"] = df_filtro[col_nota.name].values if col_nota.name != "ND" else "N/A"
                            df_export["Produto"] = df_filtro[col_prod.name].values if col_prod.name != "ND" else "N/A"
                            df_export["Descrição"] = df_filtro['Descricao'].values
                            df_export["NCM"] = df_filtro['NCM'].values
                            df_export["CST"] = df_filtro[col_cst_imposto].values
                            df_export["Base Calc"] = df_filtro[f'Base_{sufixo_imposto}'].values
                            df_export["Alíquota"] = df_filtro[f'Alq_{sufixo_imposto}'].values
                            df_export["Valor Imposto"] = df_filtro[f'Vlr_{sufixo_imposto}'].values
                            
                            excel_detalhe = exportar_excel_formatado(df_export, f"Detalhe_{sufixo_imposto}")
                            st.download_button(
                                label="📥 Excel",
                                data=excel_detalhe,
                                file_name=f"Detalhe_Saida_{sufixo_imposto}_{datetime.now().strftime('%d%m%Y')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"download_detalhe_saida_{sufixo_imposto}"
                            )

                if visao == "🟦 ICMS Normal": renderizar_drill_tributario("ICMS Normal", "ICMS")
                elif visao == "🟨 PIS": renderizar_drill_tributario("PIS", "PIS")
                elif visao == "🟩 COFINS": renderizar_drill_tributario("COFINS", "COFINS")

# 
# 📄 MÓDULO DE AUDITORIA DE NOTAS
# 

def app_auditoria_notas():
    st.header("📄 Auditoria de Notas: Dominio vs DTE")
    
    st.subheader("🔍 Auditoria NFe")
    col1, col2 = st.columns(2)
    with col1: 
        f_dominio = st.file_uploader("Dominio", type=['csv', 'xlsx', 'xls'], key="dom_up")
    with col2: 
        f_dte = st.file_uploader("DTE (Relatório Original)", type=['csv', 'xlsx', 'xls'], key="dte_up")
    
    tolerancia = st.number_input("Tolerância (R$)", value=0.05, step=0.01)
    
    if f_dominio and f_dte:
        with st.spinner('Cruzando dados...'):
            df_dom_raw, df_dte_raw = load_data(f_dominio), load_data(f_dte)
            if df_dom_raw is not None and df_dte_raw is not None:
                df_dom = localizar_cabecalho_real(df_dom_raw, ['nota', 'cont'])
                df_dte = localizar_cabecalho_real(df_dte_raw, ['nf-e', 'valor'], ['nfe', 'valor'])
                
                c_nota_dom = pegar_coluna_qualquer_chave(df_dom, ['nota', 'nf', 'documento'])
                c_val_dom = pegar_coluna_qualquer_chave(df_dom, ['cont', 'contábil', 'valor'])
                c_nota_dte = pegar_coluna_qualquer_chave(df_dte, ['nf-e', 'nfe', 'nota'])
                c_val_dte = pegar_coluna_qualquer_chave(df_dte, ['valor', 'total'])
                c_chave_dte = pegar_coluna_qualquer_chave(df_dte, ['chave', 'acesso'])
                
                if c_nota_dom.name == "ND" or c_nota_dte.name == "ND":
                    st.error("❌ Colunas não localizadas.")
                    return
                
                df_dom['Chave_Nota'] = c_nota_dom.apply(limpa_nota)
                df_dom = df_dom[df_dom['Chave_Nota'].str.contains(r'\d', na=False, regex=True)].copy()
                df_dom['Valor_Dominio'] = c_val_dom.apply(converter_valor_br)
                df_dom_grouped = df_dom.groupby('Chave_Nota')['Valor_Dominio'].sum().reset_index()
                
                df_dte['Chave_Nota'] = c_nota_dte.apply(limpa_nota)
                df_dte = df_dte[df_dte['Chave_Nota'].str.contains(r'\d', na=False, regex=True)].copy()
                df_dte['Valor_DTE'] = c_val_dte.apply(converter_valor_br)
                df_dte['Chave_Acesso_DTE'] = c_chave_dte.astype(str).replace('nan', 'Não Informada') if c_chave_dte.name != "ND" else "Não Informada"
                
                df_dte_grouped = df_dte.groupby('Chave_Nota').agg({'Valor_DTE': 'sum', 'Chave_Acesso_DTE': 'first'}).reset_index()
                
                merged = pd.merge(df_dom_grouped, df_dte_grouped, on='Chave_Nota', how='outer').fillna({'Valor_Dominio':0, 'Valor_DTE':0, 'Chave_Acesso_DTE':'N/A'})
                merged['Diferença_Bruta'] = merged['Valor_Dominio'] - merged['Valor_DTE']
                merged['Diferença_Ajustada'] = merged.apply(lambda r: 0.0 if abs(r['Diferença_Bruta']) <= tolerancia else r['Diferença_Bruta'], axis=1)
                
                d_val = merged[(merged['Diferença_Ajustada'] != 0) & (merged['Valor_Dominio'] != 0) & (merged['Valor_DTE'] != 0)]
                f_dte = merged[(merged['Valor_DTE'] == 0) & (merged['Valor_Dominio'] != 0)]
                f_dom = merged[(merged['Valor_Dominio'] == 0) & (merged['Valor_DTE'] != 0)]
                
                t1, t2, t3 = st.tabs([f"⚠️ Divergência ({len(d_val)})", f"❌ Faltam DTE ({len(f_dte)})", f"❗ Faltam Dominio ({len(f_dom)})"])
                
                with t1:
                    if not d_val.empty:
                        v_diff = d_val.copy().rename(columns={'Chave_Nota': 'N° Nota'})
                        v_diff_display = v_diff.copy()
                        v_diff_display['Valor_Dominio'] = v_diff_display['Valor_Dominio'].apply(formata_br)
                        v_diff_display['Valor_DTE'] = v_diff_display['Valor_DTE'].apply(formata_br)
                        v_diff_display['Diferença_Bruta'] = v_diff_display['Diferença_Bruta'].apply(formata_br)
                        st.dataframe(v_diff_display[['N° Nota', 'Valor_Dominio', 'Valor_DTE', 'Diferença_Bruta']], use_container_width=True)
                        
                        col_exp1, col_exp2 = st.columns([4, 1])
                        with col_exp2:
                            excel_div = exportar_excel_formatado(v_diff[['N° Nota', 'Valor_Dominio', 'Valor_DTE', 'Diferença_Bruta']], "Divergências")
                            st.download_button(
                                label="📥 Excel",
                                data=excel_div,
                                file_name=f"Divergencias_DTE_{datetime.now().strftime('%d%m%Y')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="download_div_original"
                            )
                with t2:
                    if not f_dte.empty:
                        v_fdte = f_dte.copy().rename(columns={'Chave_Nota': 'N° Nota'})
                        v_fdte_display = v_fdte.copy()
                        v_fdte_display['Valor_Dominio'] = v_fdte_display['Valor_Dominio'].apply(formata_br)
                        st.dataframe(v_fdte_display[['N° Nota', 'Valor_Dominio']], use_container_width=True)
                        
                        col_exp3, col_exp4 = st.columns([4, 1])
                        with col_exp4:
                            excel_fdte = exportar_excel_formatado(v_fdte[['N° Nota', 'Valor_Dominio']], "Faltam_DTE")
                            st.download_button(
                                label="📥 Excel",
                                data=excel_fdte,
                                file_name=f"Faltam_DTE_{datetime.now().strftime('%d%m%Y')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="download_fdte_original"
                            )
                with t3:
                    if not f_dom.empty:
                        v_fdom = f_dom.copy().rename(columns={'Chave_Nota': 'N° Nota', 'Chave_Acesso_DTE': 'Chave da Nota'})
                        v_fdom_display = v_fdom.copy()
                        v_fdom_display['Valor_DTE'] = v_fdom_display['Valor_DTE'].apply(formata_br)
                        st.dataframe(v_fdom_display[['N° Nota', 'Chave da Nota', 'Valor_DTE']], use_container_width=True)
                        
                        col_exp5, col_exp6 = st.columns([4, 1])
                        with col_exp6:
                            excel_fdom = exportar_excel_formatado(v_fdom[['N° Nota', 'Chave da Nota', 'Valor_DTE']], "Faltam_Dominio")
                            st.download_button(
                                label="📥 Excel",
                                data=excel_fdom,
                                file_name=f"Faltam_Dominio_{datetime.now().strftime('%d%m%Y')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="download_fdom_original"
                            )
    
    st.divider()
    
    st.subheader("🔍 Auditoria NFCe")
    col3, col4 = st.columns(2)
    with col3: 
        f_dominio_novo = st.file_uploader("Dominio (novo DTE)", type=['csv', 'xlsx', 'xls'], key="dom_up_novo")
    with col4: 
        f_dte_novo = st.file_uploader("DTE (Novo)", type=['csv', 'xlsx', 'xls'], key="dte_up_novo")
    
    tolerancia_novo = st.number_input("Tolerância (R$) - Novo DTE", value=0.05, step=0.01, key="tol_novo")
    
    if f_dominio_novo and f_dte_novo:
        with st.spinner('Cruzando dados...'):
            df_dom_novo_raw, df_dte_novo_raw = load_data(f_dominio_novo), load_data(f_dte_novo)
            if df_dom_novo_raw is not None and df_dte_novo_raw is not None:
                df_dom_novo = localizar_cabecalho_real(df_dom_novo_raw, ['nota', 'cont'])
                df_dte_novo = localizar_cabecalho_real(df_dte_novo_raw, ['número do documento fiscal', 'valor total'], ['numero documento', 'valor'])
                
                c_nota_dom_novo = pegar_coluna_qualquer_chave(df_dom_novo, ['nota', 'nf', 'documento'])
                c_val_dom_novo = pegar_coluna_qualquer_chave(df_dom_novo, ['cont', 'contábil', 'valor'])
                c_nota_dte_novo = pegar_coluna_qualquer_chave(df_dte_novo, ['número do documento fiscal', 'numero documento', 'nf-e', 'nfe', 'nota'])
                c_val_dte_novo = pegar_coluna_qualquer_chave(df_dte_novo, ['valor total', 'valor', 'total'])
                
                if c_nota_dom_novo.name == "ND" or c_nota_dte_novo.name == "ND":
                    st.error("❌ Colunas não localizadas.")
                    return
                
                df_dom_novo['Chave_Nota'] = c_nota_dom_novo.apply(limpa_nota)
                df_dom_novo = df_dom_novo[df_dom_novo['Chave_Nota'].str.contains(r'\d', na=False, regex=True)].copy()
                df_dom_novo['Valor_Dominio'] = c_val_dom_novo.apply(converter_valor_br)
                df_dom_novo_grouped = df_dom_novo.groupby('Chave_Nota')['Valor_Dominio'].sum().reset_index()
                
                df_dte_novo['Chave_Nota'] = c_nota_dte_novo.astype(str).str.replace('.', '', regex=False).str.strip()
                df_dte_novo = df_dte_novo[df_dte_novo['Chave_Nota'].str.contains(r'\d', na=False, regex=True)].copy()
                df_dte_novo['Valor_DTE'] = c_val_dte_novo.apply(converter_valor_br)
                
                df_dte_novo_grouped = df_dte_novo.groupby('Chave_Nota')['Valor_DTE'].sum().reset_index()
                
                merged_novo = pd.merge(df_dom_novo_grouped, df_dte_novo_grouped, on='Chave_Nota', how='outer').fillna({'Valor_Dominio':0, 'Valor_DTE':0})
                merged_novo['Diferença_Bruta'] = merged_novo['Valor_Dominio'] - merged_novo['Valor_DTE']
                merged_novo['Diferença_Ajustada'] = merged_novo.apply(lambda r: 0.0 if abs(r['Diferença_Bruta']) <= tolerancia_novo else r['Diferença_Bruta'], axis=1)
                
                d_val_novo = merged_novo[(merged_novo['Diferença_Ajustada'] != 0) & (merged_novo['Valor_Dominio'] != 0) & (merged_novo['Valor_DTE'] != 0)]
                f_dte_novo = merged_novo[(merged_novo['Valor_DTE'] == 0) & (merged_novo['Valor_Dominio'] != 0)]
                f_dom_novo = merged_novo[(merged_novo['Valor_Dominio'] == 0) & (merged_novo['Valor_DTE'] != 0)]
                
                t4, t5, t6 = st.tabs([f"⚠️ Divergência ({len(d_val_novo)})", f"❌ Faltam DTE ({len(f_dte_novo)})", f"❗ Faltam Dominio ({len(f_dom_novo)})"])
                
                with t4:
                    if not d_val_novo.empty:
                        v_diff_novo = d_val_novo.copy().rename(columns={'Chave_Nota': 'N° Nota'})
                        v_diff_novo_display = v_diff_novo.copy()
                        v_diff_novo_display['Valor_Dominio'] = v_diff_novo_display['Valor_Dominio'].apply(formata_br)
                        v_diff_novo_display['Valor_DTE'] = v_diff_novo_display['Valor_DTE'].apply(formata_br)
                        v_diff_novo_display['Diferença_Bruta'] = v_diff_novo_display['Diferença_Bruta'].apply(formata_br)
                        st.dataframe(v_diff_novo_display[['N° Nota', 'Valor_Dominio', 'Valor_DTE', 'Diferença_Bruta']], use_container_width=True)
                        
                        col_exp7, col_exp8 = st.columns([4, 1])
                        with col_exp8:
                            excel_div_novo = exportar_excel_formatado(v_diff_novo[['N° Nota', 'Valor_Dominio', 'Valor_DTE', 'Diferença_Bruta']], "Divergências")
                            st.download_button(
                                label="📥 Excel",
                                data=excel_div_novo,
                                file_name=f"Divergencias_Novo_{datetime.now().strftime('%d%m%Y')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="download_div_novo"
                            )
                with t5:
                    if not f_dte_novo.empty:
                        v_fdte_novo = f_dte_novo.copy().rename(columns={'Chave_Nota': 'N° Nota'})
                        v_fdte_novo_display = v_fdte_novo.copy()
                        v_fdte_novo_display['Valor_Dominio'] = v_fdte_novo_display['Valor_Dominio'].apply(formata_br)
                        st.dataframe(v_fdte_novo_display[['N° Nota', 'Valor_Dominio']], use_container_width=True)
                        
                        col_exp9, col_exp10 = st.columns([4, 1])
                        with col_exp10:
                            excel_fdte_novo = exportar_excel_formatado(v_fdte_novo[['N° Nota', 'Valor_Dominio']], "Faltam_DTE")
                            st.download_button(
                                label="📥 Excel",
                                data=excel_fdte_novo,
                                file_name=f"Faltam_DTE_Novo_{datetime.now().strftime('%d%m%Y')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="download_fdte_novo"
                            )
                with t6:
                    if not f_dom_novo.empty:
                        v_fdom_novo = f_dom_novo.copy().rename(columns={'Chave_Nota': 'N° Nota'})
                        v_fdom_novo_display = v_fdom_novo.copy()
                        v_fdom_novo_display['Valor_DTE'] = v_fdom_novo_display['Valor_DTE'].apply(formata_br)
                        st.dataframe(v_fdom_novo_display[['N° Nota', 'Valor_DTE']], use_container_width=True)
                        
                        col_exp11, col_exp12 = st.columns([4, 1])
                        with col_exp12:
                            excel_fdom_novo = exportar_excel_formatado(v_fdom_novo[['N° Nota', 'Valor_DTE']], "Faltam_Dominio")
                            st.download_button(
                                label="📥 Excel",
                                data=excel_fdom_novo,
                                file_name=f"Faltam_Dominio_Novo_{datetime.now().strftime('%d%m%Y')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="download_fdom_novo"
                            )

# 
# 👥 PAINEL DE GERENTE
# 

def app_painel_gerente():
    st.header("👥 Gerenciamento de Usuários")
    
    if st.session_state.tipo_usuario != 'gerente':
        st.error("❌ Acesso restrito!")
        return
    
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "Criar Usuário", 
        "Meus Usuários", 
        "Alterar Minha Senha", 
        "🔄 Bloquear/Desbloquear",
        "🔐 Alterar Senha de Subordinado"
    ])
    
    with tab1:
        st.subheader("Criar Novo Usuário Final")
        st.info(f"ℹ️ CNPJ/CPF: **{st.session_state.cnpj_cpf}**")
        
        with st.form("form_usuario_final"):
            login = st.text_input("Login")
            senha = st.text_input("Senha", type="password")
            nome = st.text_input("Nome Completo")
            email = st.text_input("Email")
            perfil = st.selectbox("Perfil", ["analista", "supervisor", "auditor"])
            
            if st.form_submit_button("Criar", use_container_width=True):
                if not all([login, senha, nome, email]):
                    st.error("❌ Campos obrigatórios!")
                elif len(senha) < 6:
                    st.error("❌ Senha com mínimo 6 caracteres!")
                else:
                    conn = sqlite3.connect(DB_PATH)
                    c = conn.cursor()
                    
                    c.execute('SELECT id FROM usuarios_finais WHERE login = ?', (login,))
                    if c.fetchone():
                        st.error(f"❌ Login já existe!")
                        conn.close()
                    else:
                        try:
                            senha_hash = hashlib.sha256(senha.encode()).hexdigest()
                            c.execute('''
                            INSERT INTO usuarios_finais 
                            (gerente_id, cnpj_cpf, login, senha_hash, nome, email, perfil, ativo)
                            VALUES (?, ?, ?, ?, ?, ?, ?, 1)
                            ''', (st.session_state.usuario_id, st.session_state.cnpj_cpf, login, 
                                 senha_hash, nome, email, perfil))
                            
                            conn.commit()
                            st.toast(f"✅ Usuário '{login}' criado!", icon="✅")
                            registrar_log(st.session_state.usuario_id, 'gerente', 
                                        st.session_state.cnpj_cpf, 'CRIACAO_USUARIO_FINAL',
                                        f"Usuário {login} com perfil {perfil}")
                        except Exception as e:
                            st.error(f"❌ Erro: {e}")
                        finally:
                            conn.close()
    
    with tab2:
        st.subheader("Meus Usuários")
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('''
        SELECT id, login, nome, email, perfil, criado_em, ativo
        FROM usuarios_finais
        WHERE gerente_id = ? AND cnpj_cpf = ?
        ORDER BY criado_em DESC
        ''', (st.session_state.usuario_id, st.session_state.cnpj_cpf))
        
        usuarios = c.fetchall()
        conn.close()
        
        if usuarios:
            df_usuarios = pd.DataFrame(usuarios, columns=[
                'ID', 'Login', 'Nome', 'Email', 'Perfil', 'Criado', 'Ativo'
            ])
            st.dataframe(df_usuarios, use_container_width=True)
        else:
            st.info("Nenhum usuário criado")
    
    with tab3:
        st.subheader("Alterar Minha Senha")
        
        with st.form("form_alterar_senha"):
            senha_atual = st.text_input("Senha Atual", type="password")
            nova_senha = st.text_input("Nova Senha", type="password")
            confirmar_senha = st.text_input("Confirmar", type="password")
            
            if st.form_submit_button("Alterar", use_container_width=True):
                conn = sqlite3.connect(DB_PATH)
                c = conn.cursor()
                
                senha_atual_hash = hashlib.sha256(senha_atual.encode()).hexdigest()
                c.execute('SELECT id FROM gerentes WHERE id = ? AND senha_hash = ?', 
                         (st.session_state.usuario_id, senha_atual_hash))
                
                if not c.fetchone():
                    st.error("❌ Senha atual incorreta!")
                elif len(nova_senha) < 6:
                    st.error("❌ Nova senha com mínimo 6 caracteres!")
                elif nova_senha != confirmar_senha:
                    st.error("❌ Senhas não conferem!")
                else:
                    nova_senha_hash = hashlib.sha256(nova_senha.encode()).hexdigest()
                    c.execute('UPDATE gerentes SET senha_hash = ? WHERE id = ?', 
                             (nova_senha_hash, st.session_state.usuario_id))
                    conn.commit()
                    st.toast("✅ Senha alterada!", icon="✅")
                    registrar_log(st.session_state.usuario_id, 'gerente', 
                                st.session_state.cnpj_cpf, 'SENHA_ALTERADA',
                                'Gerente alterou sua senha')
                
                conn.close()
    
    with tab4:
        st.subheader("🔄 Bloquear/Desbloquear Usuários Subordinados")
        
        conn = sqlite3.connect(DB_PATH)
        df_usuarios = pd.read_sql_query(
            "SELECT id, login, nome, ativo FROM usuarios_finais WHERE gerente_id = ? ORDER BY login",
            conn, params=(st.session_state.usuario_id,))
        conn.close()
        
        if not df_usuarios.empty:
            opcoes = [f"{row['login']} ({row['nome']}) - {'✅ Ativo' if row['ativo'] == 1 else '❌ Bloqueado'}" 
                     for _, row in df_usuarios.iterrows()]
            idx = st.selectbox("Selecione um usuário:", range(len(opcoes)), 
                             format_func=lambda x: opcoes[x], key="sel_bloqueio_subordinado")
            
            selecionado = df_usuarios.iloc[idx]
            usuario_id = int(selecionado['id'])
            ativo = int(selecionado['ativo'])
            login = selecionado['login']
            
            col1, col2 = st.columns(2)
            
            with col1:
                if st.button("🔒 Bloquear", use_container_width=True, key=f"bloq_sub_{usuario_id}"):
                    if ativo == 1:
                        sucesso, mensagem = bloquear_usuario_subordinado(st.session_state.usuario_id, usuario_id)
                        if sucesso:
                            st.toast(mensagem, icon="🔒")
                            
                        else:
                            st.error(mensagem)
                    else:
                        st.warning("⚠️ Já está bloqueado!")
            
            with col2:
                if st.button("🔓 Desbloquear", use_container_width=True, key=f"debloq_sub_{usuario_id}"):
                    if ativo == 0:
                        sucesso, mensagem = desbloquear_usuario_subordinado(st.session_state.usuario_id, usuario_id)
                        if sucesso:
                            st.toast(mensagem, icon="🔓")
                            
                        else:
                            st.error(mensagem)
                    else:
                        st.warning("⚠️ Já está desbloqueado!")
        else:
            st.info("Você não possui usuários subordinados")
    
    with tab5:
        st.subheader("🔐 Alterar Senha de Subordinado")
        st.info("💡 Permissão para redefinir senhas de usuários bloqueados ou que esqueceram a senha")
        
        conn = sqlite3.connect(DB_PATH)
        df_usuarios = pd.read_sql_query(
            "SELECT id, login, nome, ativo FROM usuarios_finais WHERE gerente_id = ? ORDER BY login",
            conn, params=(st.session_state.usuario_id,))
        conn.close()
        
        if not df_usuarios.empty:
            opcoes = [f"{row['login']} ({row['nome']}) - {'✅ Ativo' if row['ativo'] == 1 else '❌ Bloqueado'}" 
                     for _, row in df_usuarios.iterrows()]
            idx = st.selectbox("Selecione um subordinado:", range(len(opcoes)), 
                             format_func=lambda x: opcoes[x], key="sel_alterar_senha_subordinado")
            
            selecionado = df_usuarios.iloc[idx]
            usuario_id = int(selecionado['id'])
            login = selecionado['login']
            
            st.warning(f"🔐 Redefinir senha de: **{login}**")
            
            with st.form("form_alterar_senha_subordinado"):
                nova_senha = st.text_input("Nova Senha", type="password", key="nova_senha_sub")
                confirmar_senha = st.text_input("Confirmar Senha", type="password", key="conf_senha_sub")
                
                if st.form_submit_button("🔐 Redefinir Senha", use_container_width=True):
                    if len(nova_senha) < 6:
                        st.error("❌ Senha com mínimo 6 caracteres!")
                    elif nova_senha != confirmar_senha:
                        st.error("❌ Senhas não conferem!")
                    else:
                        try:
                            nova_senha_hash = hashlib.sha256(nova_senha.encode()).hexdigest()
                            
                            conn = sqlite3.connect(DB_PATH)
                            c = conn.cursor()
                            c.execute('UPDATE usuarios_finais SET senha_hash = ? WHERE id = ?', 
                                     (nova_senha_hash, usuario_id))
                            conn.commit()
                            conn.close()
                            
                            st.toast(f"✅ Senha de '{login}' redefinida com sucesso!", icon="✅")
                            registrar_log(st.session_state.usuario_id, 'gerente', 
                                        st.session_state.cnpj_cpf, 'SENHA_SUBORDINADO_ALTERADA',
                                        f'Gerente alterou senha de {login} (ID {usuario_id})')
                            
                        except Exception as e:
                            st.error(f"❌ Erro: {str(e)}")
        else:
            st.info("Você não possui usuários subordinados")

# 
# ⚙️ PAINEL ADMIN
# 

def app_painel_admin():
    st.header("⚙️ Painel de Administração")
    
    if st.session_state.tipo_usuario != 'admin':
        st.error("❌ Acesso restrito!")
        return
    
    tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
        "👥 Gerentes", "👤 Usuários", "🔑 Senha", "🔄 Bloqueio", 
        "🗑️ Excluir", "💰 Financeiro", "🛡️ Logs"
    ])
    
    with tab1:
        st.subheader("Criar Novo Gerente")
        with st.form("form_gerente"):
            login = st.text_input("Login")
            senha = st.text_input("Senha", type="password")
            cnpj_cpf = st.text_input("CNPJ/CPF")
            nome_empresa = st.text_input("Empresa")
            email = st.text_input("Email")
            
            if st.form_submit_button("Criar", use_container_width=True):
                if not all([login, senha, cnpj_cpf, nome_empresa]):
                    st.error("❌ Campos obrigatórios!")
                else:
                    conn = sqlite3.connect(DB_PATH)
                    c = conn.cursor()
                    
                    c.execute('SELECT id FROM gerentes WHERE cnpj_cpf = ?', (cnpj_cpf,))
                    if c.fetchone():
                        st.error(f"❌ CNPJ/CPF já cadastrado!")
                        conn.close()
                    else:
                        senha_hash = hashlib.sha256(senha.encode()).hexdigest()
                        try:
                            c.execute('''
                            INSERT INTO gerentes (admin_id, login, senha_hash, cnpj_cpf, nome_empresa, email, ativo)
                            VALUES (?, ?, ?, ?, ?, ?, 1)
                            ''', (st.session_state.usuario_id, login, senha_hash, cnpj_cpf, nome_empresa, email))
                            
                            conn.commit()
                            st.toast(f"✅ Gerente '{login}' criado!", icon="✅")
                            registrar_log(st.session_state.usuario_id, 'admin', '', 
                                        'CRIACAO_GERENTE', f"Gerente {login} para {nome_empresa}")
                        except Exception as e:
                            st.error(f"❌ Erro: {e}")
                        finally:
                            conn.close()
        
        st.divider()
        st.subheader("Gerentes Cadastrados")
        conn = sqlite3.connect(DB_PATH)
        df_gerentes = pd.read_sql_query(
            "SELECT id, login, cnpj_cpf, nome_empresa, email, criado_em, ativo FROM gerentes ORDER BY criado_em DESC",
            conn
        )
        conn.close()
        
        if not df_gerentes.empty:
            st.dataframe(df_gerentes, use_container_width=True)
        else:
            st.info("Nenhum gerente cadastrado")
    
    with tab2:
        st.subheader("Usuários Finais")
        
        filtro_status = st.selectbox("Filtro:", ["Todos", "Ativos", "Bloqueados"], key="filtro_usuarios")
        
        conn = sqlite3.connect(DB_PATH)
        
        if filtro_status == "Todos":
            df_usuarios = pd.read_sql_query(
                "SELECT id, login, cnpj_cpf, nome, perfil, criado_em, ativo FROM usuarios_finais ORDER BY criado_em DESC",
                conn
            )
        elif filtro_status == "Ativos":
            df_usuarios = pd.read_sql_query(
                "SELECT id, login, cnpj_cpf, nome, perfil, criado_em, ativo FROM usuarios_finais WHERE ativo = 1 ORDER BY criado_em DESC",
                conn
            )
        else:
            df_usuarios = pd.read_sql_query(
                "SELECT id, login, cnpj_cpf, nome, perfil, criado_em, ativo FROM usuarios_finais WHERE ativo = 0 ORDER BY criado_em DESC",
                conn
            )
        
        conn.close()
        
        if not df_usuarios.empty:
            st.dataframe(df_usuarios, use_container_width=True)
        else:
            st.info(f"Nenhum usuário {filtro_status.lower()}")
    
    with tab3:
        st.subheader("Alterar Senha")
        
        tipo_alvo = st.radio("Tipo:", ["Gerente", "Usuário Final"], key="tipo_senha")
        conn = sqlite3.connect(DB_PATH)
        
        if tipo_alvo == "Gerente":
            df_gerentes = pd.read_sql_query("SELECT id, login, nome_empresa FROM gerentes", conn)
            
            if not df_gerentes.empty:
                usuario = st.selectbox("Selecione:", 
                    [f"{row['login']} ({row['nome_empresa']})" for _, row in df_gerentes.iterrows()],
                    key="sel_gerente_senha")
                
                usuario_id = int(df_gerentes[df_gerentes['login'] == usuario.split(' (')[0]]['id'].values[0])
                
                with st.form("form_senha_gerente"):
                    nova = st.text_input("Nova Senha", type="password", key="nova_gerente")
                    conf = st.text_input("Confirmar", type="password", key="conf_gerente")
                    
                    if st.form_submit_button("Alterar", use_container_width=True):
                        if len(nova) < 6:
                            st.error("❌ Mínimo 6 caracteres!")
                        elif nova != conf:
                            st.error("❌ Senhas não conferem!")
                        else:
                            hash_nova = hashlib.sha256(nova.encode()).hexdigest()
                            c = conn.cursor()
                            c.execute('UPDATE gerentes SET senha_hash = ? WHERE id = ?', (hash_nova, usuario_id))
                            conn.commit()
                            st.toast("✅ Senha alterada!", icon="✅")
                            registrar_log(st.session_state.usuario_id, 'admin', '', 'SENHA_ALTERADA',
                                        f"Senha de gerente ID {usuario_id}")
        else:
            df_usuarios = pd.read_sql_query("SELECT id, login, nome FROM usuarios_finais", conn)
            
            if not df_usuarios.empty:
                usuario = st.selectbox("Selecione:",
                    [f"{row['login']} ({row['nome']})" for _, row in df_usuarios.iterrows()],
                    key="sel_usuario_senha")
                
                usuario_id = int(df_usuarios[df_usuarios['login'] == usuario.split(' (')[0]]['id'].values[0])
                
                with st.form("form_senha_usuario"):
                    nova = st.text_input("Nova Senha", type="password", key="nova_usuario")
                    conf = st.text_input("Confirmar", type="password", key="conf_usuario")
                    
                    if st.form_submit_button("Alterar", use_container_width=True):
                        if len(nova) < 6:
                            st.error("❌ Mínimo 6 caracteres!")
                        elif nova != conf:
                            st.error("❌ Senhas não conferem!")
                        else:
                            hash_nova = hashlib.sha256(nova.encode()).hexdigest()
                            c = conn.cursor()
                            c.execute('UPDATE usuarios_finais SET senha_hash = ? WHERE id = ?', (hash_nova, usuario_id))
                            conn.commit()
                            st.toast("✅ Senha alterada!", icon="✅")
                            registrar_log(st.session_state.usuario_id, 'admin', '', 'SENHA_ALTERADA',
                                        f"Senha de usuário ID {usuario_id}")
        
        conn.close()
    
    with tab4:
        st.subheader("🔄 Bloquear/Desbloquear")
        
        tipo_bloqueio = st.radio("Tipo:", ["Gerente", "Usuário Final"], key="tipo_bloqueio")
        
        if tipo_bloqueio == "Gerente":
            conn = sqlite3.connect(DB_PATH)
            df_gerentes = pd.read_sql_query(
                "SELECT id, login, nome_empresa, ativo FROM gerentes ORDER BY login", conn)
            conn.close()
            
            if not df_gerentes.empty:
                opcoes = [f"{row['login']} ({row['nome_empresa']}) - {'✅ Ativo' if row['ativo'] == 1 else '❌ Bloqueado'}" 
                         for _, row in df_gerentes.iterrows()]
                idx = st.selectbox("Selecione:", range(len(opcoes)), 
                                 format_func=lambda x: opcoes[x], key="sel_bloqueio_gerente")
                
                selecionado = df_gerentes.iloc[idx]
                usuario_id = int(selecionado['id'])
                ativo = int(selecionado['ativo'])
                login = selecionado['login']
                
                col1, col2 = st.columns(2)
                
                with col1:
                    if st.button("🔒 Bloquear", use_container_width=True, key=f"bloq_gerente_{usuario_id}"):
                        if ativo == 1:
                            conn = sqlite3.connect(DB_PATH)
                            c = conn.cursor()
                            c.execute('UPDATE gerentes SET ativo = 0 WHERE id = ?', (usuario_id,))
                            conn.commit()
                            conn.close()
                            st.toast(f"✅ Gerente '{login}' bloqueado!", icon="🔒")
                            registrar_log(st.session_state.usuario_id, 'admin', '', 'GERENTE_BLOQUEADO',
                                        f"Gerente {login} (ID {usuario_id}) bloqueado")
                            
                        else:
                            st.warning("⚠️ Já está bloqueado!")
                
                with col2:
                    if st.button("🔓 Desbloquear", use_container_width=True, key=f"debloq_gerente_{usuario_id}"):
                        if ativo == 0:
                            conn = sqlite3.connect(DB_PATH)
                            c = conn.cursor()
                            c.execute('UPDATE gerentes SET ativo = 1 WHERE id = ?', (usuario_id,))
                            conn.commit()
                            conn.close()
                            st.toast(f"✅ Gerente '{login}' desbloqueado!", icon="🔓")
                            registrar_log(st.session_state.usuario_id, 'admin', '', 'GERENTE_DESBLOQUEADO',
                                        f"Gerente {login} (ID {usuario_id}) desbloqueado")
                            
                        else:
                            st.warning("⚠️ Já está desbloqueado!")
        
        else:
            conn = sqlite3.connect(DB_PATH)
            df_usuarios = pd.read_sql_query(
                "SELECT id, login, nome, ativo FROM usuarios_finais ORDER BY login", conn)
            conn.close()
            
            if not df_usuarios.empty:
                opcoes = [f"{row['login']} ({row['nome']}) - {'✅ Ativo' if row['ativo'] == 1 else '❌ Bloqueado'}" 
                         for _, row in df_usuarios.iterrows()]
                idx = st.selectbox("Selecione:", range(len(opcoes)), 
                                 format_func=lambda x: opcoes[x], key="sel_bloqueio_usuario")
                
                selecionado = df_usuarios.iloc[idx]
                usuario_id = int(selecionado['id'])
                ativo = int(selecionado['ativo'])
                login = selecionado['login']
                
                col1, col2 = st.columns(2)
                
                with col1:
                    if st.button("🔒 Bloquear", use_container_width=True, key=f"bloq_usuario_{usuario_id}"):
                        if ativo == 1:
                            conn = sqlite3.connect(DB_PATH)
                            c = conn.cursor()
                            c.execute('UPDATE usuarios_finais SET ativo = 0 WHERE id = ?', (usuario_id,))
                            conn.commit()
                            conn.close()
                            st.toast(f"✅ Usuário '{login}' bloqueado!", icon="🔒")
                            registrar_log(st.session_state.usuario_id, 'admin', '', 'USUARIO_BLOQUEADO',
                                        f"Usuário {login} (ID {usuario_id}) bloqueado")
                            
                        else:
                            st.warning("⚠️ Já está bloqueado!")
                
                with col2:
                    if st.button("🔓 Desbloquear", use_container_width=True, key=f"debloq_usuario_{usuario_id}"):
                        if ativo == 0:
                            conn = sqlite3.connect(DB_PATH)
                            c = conn.cursor()
                            c.execute('UPDATE usuarios_finais SET ativo = 1 WHERE id = ?', (usuario_id,))
                            conn.commit()
                            conn.close()
                            st.toast(f"✅ Usuário '{login}' desbloqueado!", icon="🔓")
                            registrar_log(st.session_state.usuario_id, 'admin', '', 'USUARIO_DESBLOQUEADO',
                                        f"Usuário {login} (ID {usuario_id}) desbloqueado")
                            
                        else:
                            st.warning("⚠️ Já está desbloqueado!")
    
    with tab5:
        st.subheader("🗑️ Excluir - IRREVERSÍVEL")
        st.warning("⚠️ Exclusão permanente!")
        
        tipo_excluir = st.radio("Tipo:", ["Gerente", "Usuário Final"], key="tipo_excluir")
        
        if tipo_excluir == "Gerente":
            conn = sqlite3.connect(DB_PATH)
            df_gerentes = pd.read_sql_query("SELECT id, login, nome_empresa FROM gerentes ORDER BY login", conn)
            conn.close()
            
            if not df_gerentes.empty:
                opcoes = [f"{row['login']} ({row['nome_empresa']})" for _, row in df_gerentes.iterrows()]
                idx = st.selectbox("Selecione:", range(len(opcoes)), 
                                 format_func=lambda x: opcoes[x], key="sel_excluir_gerente")
                
                selecionado = df_gerentes.iloc[idx]
                usuario_id = int(selecionado['id'])
                login = selecionado['login']
                empresa = selecionado['nome_empresa']
                
                st.error(f"🗑️ Deletar: **{login}** ({empresa})?")
                
                conf = st.text_input(f"Digite '{login}' para confirmar:", key="conf_excluir_gerente")
                
                if st.button("🗑️ EXCLUIR PERMANENTEMENTE", use_container_width=True, key=f"exec_excluir_gerente_{usuario_id}"):
                    if conf == login:
                        conn = sqlite3.connect(DB_PATH)
                        c = conn.cursor()
                        c.execute('DELETE FROM usuarios_finais WHERE gerente_id = ?', (usuario_id,))
                        c.execute('DELETE FROM gerentes WHERE id = ?', (usuario_id,))
                        conn.commit()
                        conn.close()
                        
                        st.toast(f"✅ Gerente '{login}' e subordinados excluídos!", icon="🗑️")
                        registrar_log(st.session_state.usuario_id, 'admin', '', 'GERENTE_EXCLUIDO',
                                    f"Gerente {login} (ID {usuario_id}) excluído permanentemente")
                        
                    else:
                        st.error("❌ Confirmação incorreta!")
        
        else:
            conn = sqlite3.connect(DB_PATH)
            df_usuarios = pd.read_sql_query("SELECT id, login, nome FROM usuarios_finais ORDER BY login", conn)
            conn.close()
            
            if not df_usuarios.empty:
                opcoes = [f"{row['login']} ({row['nome']})" for _, row in df_usuarios.iterrows()]
                idx = st.selectbox("Selecione:", range(len(opcoes)), 
                                 format_func=lambda x: opcoes[x], key="sel_excluir_usuario")
                
                selecionado = df_usuarios.iloc[idx]
                usuario_id = int(selecionado['id'])
                login = selecionado['login']
                nome = selecionado['nome']
                
                st.error(f"🗑️ Deletar: **{login}** ({nome})?")
                
                conf = st.text_input(f"Digite '{login}' para confirmar:", key="conf_excluir_usuario")
                
                if st.button("🗑️ EXCLUIR PERMANENTEMENTE", use_container_width=True, key=f"exec_excluir_usuario_{usuario_id}"):
                    if conf == login:
                        conn = sqlite3.connect(DB_PATH)
                        c = conn.cursor()
                        c.execute('DELETE FROM usuarios_finais WHERE id = ?', (usuario_id,))
                        conn.commit()
                        conn.close()
                        
                        st.toast(f"✅ Usuário '{login}' excluído!", icon="🗑️")
                        registrar_log(st.session_state.usuario_id, 'admin', '', 'USUARIO_EXCLUIDO',
                                    f"Usuário {login} (ID {usuario_id}) excluído permanentemente")
                        
                    else:
                        st.error("❌ Confirmação incorreta!")
    
    with tab6:
        st.subheader("💰 Controle Financeiro & 🆘 Suporte")
        
        subtab_fin, subtab_suporte = st.tabs(["💰 Financeiro", "🆘 Suporte"])
        
        with subtab_fin:
            st.markdown("#### 📅 Contas Recorrentes (Fixas Mensalmente)")
            st.info("💡 Cria contas que se repetem automaticamente todo mês no mesmo dia")
            
            subtab_rec1, subtab_rec2, subtab_rec3 = st.tabs([
                "➕ Nova Recorrente", 
                "📋 Listar Recorrentes", 
                "⏸️ Desativar"
            ])
            
            with subtab_rec1:
                st.markdown("##### Criar Conta Recorrente Mensal")
                
                with st.form("form_conta_recorrente"):
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        conn = sqlite3.connect(DB_PATH)
                        c = conn.cursor()
                        c.execute('''
                        SELECT id, login, cnpj_cpf, nome_empresa 
                        FROM gerentes WHERE ativo = 1 ORDER BY login
                        ''')
                        gerentes = c.fetchall()
                        conn.close()
                        
                        if gerentes:
                            gerente_opcoes = [f"{g[1]} ({g[3]})" for g in gerentes]
                            gerente_selecionado = st.selectbox(
                                "Selecione o Gerente:",
                                gerente_opcoes,
                                key="sel_gerente_recorrente"
                            )
                            gerente_id = int([g[0] for g in gerentes 
                                            if f"{g[1]} ({g[3]})" == gerente_selecionado][0])
                            gerente_cnpj = [g[2] for g in gerentes 
                                           if g[0] == gerente_id][0]
                        else:
                            st.error("❌ Nenhum gerente ativo disponível")
                            gerente_id = None
                    
                    with col2:
                        if gerente_id:
                            conn = sqlite3.connect(DB_PATH)
                            c = conn.cursor()
                            c.execute('''
                            SELECT id, login, cnpj_cpf, nome 
                            FROM usuarios_finais 
                            WHERE gerente_id = ? AND ativo = 1 
                            ORDER BY login
                            ''', (gerente_id,))
                            usuarios = c.fetchall()
                            conn.close()
                            
                            if usuarios:
                                usuario_opcoes = [f"{u[1]} ({u[2]})" for u in usuarios]
                                usuario_selecionado = st.selectbox(
                                    "Selecione o Usuário/Cliente:",
                                    usuario_opcoes,
                                    key="sel_usuario_recorrente"
                                )
                                usuario_id = int([u[0] for u in usuarios 
                                                if f"{u[1]} ({u[2]})" == usuario_selecionado][0])
                            else:
                                st.warning("⚠️ Este gerente não possui usuários vinculados")
                                usuario_id = None
                        else:
                            usuario_id = None
                    
                    col3, col4 = st.columns(2)
                    
                    with col3:
                        valor_fixo = st.number_input(
                            "Valor Fixo Mensal (R$)",
                            min_value=0.01,
                            step=0.01,
                            value=1000.00,
                            key="valor_recorrente"
                        )
                    
                    with col4:
                        dia_vencimento = st.number_input(
                            "Dia do Vencimento (1-31)",
                            min_value=1,
                            max_value=31,
                            value=10,
                            key="dia_vencimento_recorrente"
                        )
                    
                    descricao = st.text_area(
                        "Descrição (Opcional)",
                        placeholder="Ex: Consultoria contábil - Mensal",
                        key="desc_recorrente",
                        height=50
                    )
                    
                    if st.form_submit_button("✅ Criar Conta Recorrente", use_container_width=True):
                        if not gerente_id or not usuario_id:
                            st.error("❌ Selecione gerente e usuário")
                        elif valor_fixo <= 0:
                            st.error("❌ Valor deve ser maior que zero")
                        else:
                            sucesso, mensagem = criar_conta_recorrente(
                                admin_id=st.session_state.usuario_id,
                                gerente_id=gerente_id,
                                cnpj_cpf_gerente=gerente_cnpj,
                                usuario_id=usuario_id,
                                login_usuario=[u[1] for u in usuarios if u[0] == usuario_id][0],
                                cnpj_cpf_usuario=[u[2] for u in usuarios if u[0] == usuario_id][0],
                                valor_fixo=valor_fixo,
                                dia_vencimento=int(dia_vencimento),
                                descricao=descricao
                            )
                            
                            if sucesso:
                                st.toast(mensagem, icon="✅")
                                registrar_log(st.session_state.usuario_id, 'admin', '', 
                                            'CONTA_RECORRENTE_CRIADA',
                                            f'Valor: R$ {valor_fixo:.2f} - Dia: {int(dia_vencimento)}')
                                
                            else:
                                st.error(mensagem)
            
            with subtab_rec2:
                st.markdown("##### Contas Recorrentes Ativas")
                
                df_recorrentes = listar_contas_recorrentes(admin_id=st.session_state.usuario_id)
                
                if not df_recorrentes.empty:
                    df_display = df_recorrentes.copy()
                    df_display['valor_fixo'] = df_display['valor_fixo'].apply(formata_br)
                    df_display['dia_vencimento'] = df_display['dia_vencimento'].astype(str) + 'º do mês'
                    df_display.columns = ['ID', 'Usuário', 'Valor Mensal', 'Vencimento', 'Descrição', 
                                         'Gerente', 'CNPJ Gerente', 'Status']
                    
                    st.dataframe(df_display, use_container_width=True, hide_index=True)
                    
                    st.success("✅ Total de contas recorrentes ativas: " + str(len(df_recorrentes)))
                else:
                    st.info("📭 Nenhuma conta recorrente cadastrada")
            
            with subtab_rec3:
                st.markdown("##### Desativar Conta Recorrente")
                
                df_recorrentes = listar_contas_recorrentes(admin_id=st.session_state.usuario_id)
                
                if not df_recorrentes.empty:
                    opcoes = [f"{row['login_usuario']} - R$ {row['valor_fixo']:.2f}" 
                             for _, row in df_recorrentes.iterrows()]
                    
                    idx = st.selectbox("Selecione a conta a desativar:", 
                                      range(len(opcoes)),
                                      format_func=lambda x: opcoes[x],
                                      key="sel_desativar_recorrente")
                    
                    selecionada = df_recorrentes.iloc[idx]
                    recorrente_id = int(selecionada['id'])
                    
                    st.warning(f"⚠️ Desativar recorrência de {selecionada['login_usuario']}?")
                    st.info("💡 Contas já geradas não serão afetadas, apenas novas gerações serão impedidas")
                    
                    if st.button("⏸️ Desativar", use_container_width=True, key=f"desativ_{recorrente_id}"):
                        sucesso, mensagem = desativar_conta_recorrente(
                            recorrente_id=recorrente_id,
                            admin_id=st.session_state.usuario_id
                        )
                        
                        if sucesso:
                            st.toast(mensagem, icon="⏸️")
                            
                        else:
                            st.error(mensagem)
                else:
                    st.info("📭 Nenhuma conta recorrente para desativar")
        
        with subtab_suporte:
            st.subheader("🎟️ Gerenciamento de Tickets de Suporte")
            
            # Estatísticas
            stats = obter_estatisticas_tickets()
            
            col_stat1, col_stat2, col_stat3, col_stat4, col_stat5 = st.columns(5)
            
            with col_stat1:
                st.metric("📊 Total", stats['total'])
            
            with col_stat2:
                st.metric("⏳ Abertos", stats['abertos'], delta=f"🔴 {stats['criticos']} críticos" if stats['criticos'] > 0 else "✅ OK")
            
            with col_stat3:
                st.metric("✅ Respondidos", stats['respondidos'])
            
            with col_stat4:
                st.metric("🔒 Fechados", stats['fechados'])
            
            with col_stat5:
                if stats['criticos'] > 0:
                    st.error(f"🔴 {stats['criticos']} Crítico(s)")
                else:
                    st.success("✅ Sem críticos")
            
            st.divider()
            
            # Filtros
            col_filtro1, col_filtro2, col_filtro3 = st.columns(3)
            
            with col_filtro1:
                filtro_status_admin = st.multiselect(
                    "Filtrar por Status:",
                    ["ABERTO", "RESPONDIDO", "FECHADO"],
                    default=["ABERTO"],
                    key="filtro_status_admin"
                )
            
            with col_filtro2:
                filtro_prioridade = st.multiselect(
                    "Filtrar por Prioridade:",
                    ["CRITICA", "ALTA", "MEDIA", "BAIXA"],
                    default=["CRITICA", "ALTA"],
                    key="filtro_prioridade"
                )
            
            with col_filtro3:
                ordenar_por = st.selectbox(
                    "Ordenar por:",
                    ["Data (Mais Recente)", "Prioridade", "Status"],
                    key="ordenar_por"
                )
            
            st.divider()
            
            # Listar tickets
            df_tickets = listar_todos_tickets_admin()
            
            if not df_tickets.empty:
                # Aplicar filtros
                df_filtrado = df_tickets[
                    (df_tickets['status'].isin(filtro_status_admin)) &
                    (df_tickets['prioridade'].isin(filtro_prioridade))
                ]
                
                if not df_filtrado.empty:
                    st.write(f"**Total de tickets encontrados:** {len(df_filtrado)}")
                    
                    for idx, row in df_filtrado.iterrows():
                        # Emoji de status
                        if row['status'] == 'ABERTO':
                            emoji_status = "⏳"
                            cor_status = "🔴"
                        elif row['status'] == 'RESPONDIDO':
                            emoji_status = "✅"
                            cor_status = "🟢"
                        else:
                            emoji_status = "🔒"
                            cor_status = "⚪"
                        
                        # Emoji de prioridade
                        if row['prioridade'] == 'CRITICA':
                            emoji_prio = "🔴"
                        elif row['prioridade'] == 'ALTA':
                            emoji_prio = "🟠"
                        elif row['prioridade'] == 'MEDIA':
                            emoji_prio = "🟡"
                        else:
                            emoji_prio = "🟢"
                        
                        with st.expander(f"{emoji_status} #{row['id']} - {row['assunto']} {emoji_prio} - {row['usuario_login']}"):
                            col_det1, col_det2 = st.columns([2, 1])
                            
                            with col_det1:
                                st.write(f"**👤 Usuário:** {row['usuario_login']}")
                                st.write(f"**📧 E-mail:** {row['usuario_email']}")
                                st.write(f"**📝 Tipo:** {row['tipo_problema']}")
                                st.write(f"**📊 Prioridade:** {row['prioridade']}")
                                st.write(f"**📅 Criado em:** {row['data_criacao']}")
                                
                                if row['status'] == 'RESPONDIDO':
                                    st.write(f"**✅ Respondido em:** {row['data_resposta']}")
                            
                            with col_det2:
                                st.write(f"**Status:** {row['status']}")
                                if row['status'] == 'ABERTO':
                                    st.error("⏳ Pendente")
                                elif row['status'] == 'RESPONDIDO':
                                    st.info("✅ Respondido")
                                else:
                                    st.success("🔒 Fechado")
                            
                            st.divider()
                            st.markdown("**📌 Descrição do Problema:**")
                            st.info(row['descricao'])
                            
                            # Se já foi respondido
                            if row['status'] != 'ABERTO':
                                st.divider()
                                st.markdown("**💬 Resposta do Admin:**")
                                st.success(row['resposta_admin'])
                            
                            st.divider()
                            
                            # Formulário de resposta (apenas se ainda está aberto)
                            if row['status'] == 'ABERTO':
                                with st.form(f"form_resposta_{row['id']}"):
                                    resposta = st.text_area(
                                        "Sua Resposta:",
                                        height=150,
                                        placeholder="Digite a resposta para o usuário...",
                                        key=f"resposta_{row['id']}"
                                    )
                                    
                                    col_resp1, col_resp2, col_resp3 = st.columns(3)
                                    
                                    with col_resp1:
                                        if st.form_submit_button("✅ Responder", use_container_width=True, key=f"btn_resp_{row['id']}"):
                                            if resposta.strip():
                                                sucesso, msg = responder_ticket(row['id'], resposta, st.session_state.usuario_id)
                                                if sucesso:
                                                    st.success(msg)
                                                    
                                                else:
                                                    st.error(msg)
                                            else:
                                                st.error("❌ Digite uma resposta!")
                                    
                                    with col_resp2:
                                        if st.form_submit_button("🔒 Responder & Fechar", use_container_width=True, key=f"btn_resp_fechar_{row['id']}"):
                                            if resposta.strip():
                                                sucesso, msg = responder_ticket(row['id'], resposta, st.session_state.usuario_id)
                                                if sucesso:
                                                    sucesso2, msg2 = fechar_ticket(row['id'], st.session_state.usuario_id)
                                                    st.success(f"{msg}\n{msg2}")
                                                    
                                                else:
                                                    st.error(msg)
                                            else:
                                                st.error("❌ Digite uma resposta!")
                                    
                                    with col_resp3:
                                        if st.form_submit_button("❌ Fechar sem Responder", use_container_width=True, key=f"btn_fechar_{row['id']}"):
                                            sucesso, msg = fechar_ticket(row['id'], st.session_state.usuario_id)
                                            if sucesso:
                                                st.success(msg)
                                                
                                            else:
                                                st.error(msg)
                else:
                    st.info("📭 Nenhum ticket encontrado com esses filtros.")
            else:
                st.info("📭 Nenhum ticket de suporte ainda.")
    
    with tab7:
        st.subheader("🛡️ Logs (Últimos 100)")
        conn = sqlite3.connect(DB_PATH)
        df_logs = pd.read_sql_query(
            "SELECT data_hora, usuario_id, tipo_usuario, cnpj_cpf, acao, dados_novos FROM logs_auditoria ORDER BY id DESC LIMIT 100",
            conn
        )
        conn.close()
        
        if not df_logs.empty:
            st.dataframe(df_logs, use_container_width=True)
        else:
            st.info("Nenhum log")

# 
# 🆘 PAINEL DE SUPORTE PARA USUÁRIOS
# 

def app_painel_suporte_usuario():
    """Painel de suporte para usuários finais"""
    st.header("🆘 Suporte ao Sistema")
    
    st.info("💡 Abra uma solicitação de suporte e nossa equipe responderá dentro do painel. Você poderá acompanhar o status em tempo real.")
    
    tab_novo, tab_meus = st.tabs(["➕ Nova Solicitação", "📋 Meus Tickets"])
    
    with tab_novo:
        st.subheader("Abrir Nova Solicitação")
        
        with st.form("form_ticket_suporte"):
            tipo_problema = st.selectbox(
                "Tipo de Problema:",
                ["❓ Dúvida", "🐛 Bug/Erro", "🔧 Funcionalidade", "📋 Relatório", 
                 "🔐 Acesso", "⚡ Performance", "💾 Dados", "📞 Outro"]
            )
            
            assunto = st.text_input(
                "Assunto (máx. 100 caracteres):",
                max_chars=100,
                placeholder="Ex: Erro ao fazer download do relatório"
            )
            
            descricao = st.text_area(
                "Descrição Detalhada:",
                height=200,
                placeholder="Descreva o problema com detalhes:\n- O que você estava fazendo?\n- Qual é o erro exato?\n- Quando ocorreu?",
                max_chars=2000
            )
            
            col1, col2 = st.columns(2)
            
            with col1:
                prioridade = st.radio(
                    "Prioridade:",
                    ["🔴 CRÍTICA", "🟠 ALTA", "🟡 MÉDIA", "🟢 BAIXA"],
                    horizontal=False
                )
            
            with col2:
                st.info("**Prioridades:**\n\n🔴 Crítica: Sistema inoperável\n\n🟠 Alta: Funcionalidade prejudicada\n\n🟡 Média: Incômodo menor\n\n🟢 Baixa: Dúvida/Sugestão")
            
            # Mapear prioridade
            prioridade_mapa = {
                "🔴 CRÍTICA": "CRITICA",
                "🟠 ALTA": "ALTA",
                "🟡 MÉDIA": "MEDIA",
                "🟢 BAIXA": "BAIXA"
            }
            
            if st.form_submit_button("📤 Enviar Solicitação", use_container_width=True):
                if not all([assunto, descricao]):
                    st.error("❌ Preencha todos os campos obrigatórios!")
                else:
                    # Obter e-mail do usuário
                    conn = sqlite3.connect(DB_PATH)
                    c = conn.cursor()
                    c.execute('SELECT email FROM usuarios_finais WHERE id = ?', (st.session_state.usuario_id,))
                    resultado = c.fetchone()
                    email_usuario = resultado[0] if resultado else "email_nao_informado@example.com"
                    conn.close()
                    
                    sucesso, mensagem = criar_ticket_suporte(
                        usuario_id=st.session_state.usuario_id,
                        usuario_login=st.session_state.usuario,
                        usuario_email=email_usuario,
                        tipo_problema=tipo_problema,
                        assunto=assunto,
                        descricao=descricao,
                        prioridade=prioridade_mapa[prioridade]
                    )
                    
                    if sucesso:
                        st.success(mensagem)
                        st.balloons()
                    else:
                        st.error(mensagem)
    
    with tab_meus:
        st.subheader("Meus Tickets de Suporte")
        
        df_meus_tickets = listar_meus_tickets(st.session_state.usuario_id)
        
        if not df_meus_tickets.empty:
            # Filtro de status
            filtro_status = st.selectbox(
                "Filtrar por Status:",
                ["Todos", "ABERTO", "RESPONDIDO", "FECHADO"],
                key="filtro_status_usuario"
            )
            
            if filtro_status != "Todos":
                df_filtrado = df_meus_tickets[df_meus_tickets['status'] == filtro_status]
            else:
                df_filtrado = df_meus_tickets
            
            if not df_filtrado.empty:
                for idx, row in df_filtrado.iterrows():
                    # Emoji de status
                    if row['status'] == 'ABERTO':
                        emoji_status = "⏳"
                        cor = "warning"
                    elif row['status'] == 'RESPONDIDO':
                        emoji_status = "✅"
                        cor = "info"
                    else:
                        emoji_status = "🔒"
                        cor = "success"
                    
                    # Emoji de prioridade
                    if row['prioridade'] == 'CRITICA':
                        emoji_prio = "🔴"
                    elif row['prioridade'] == 'ALTA':
                        emoji_prio = "🟠"
                    elif row['prioridade'] == 'MEDIA':
                        emoji_prio = "🟡"
                    else:
                        emoji_prio = "🟢"
                    
                    with st.expander(f"{emoji_status} #{row['id']} - {row['assunto']} {emoji_prio}"):
                        st.write(f"**Tipo:** {row['tipo_problema']}")
                        st.write(f"**Status:** {row['status']}")
                        st.write(f"**Prioridade:** {row['prioridade']}")
                        st.write(f"**Criado em:** {row['data_criacao']}")
                        
                        st.divider()
                        
                        if row['status'] != 'ABERTO':
                            st.success("✅ **SUA SOLICITAÇÃO FOI RESPONDIDA**")
                            st.info(row['resposta_admin'])
            else:
                st.info(f"📭 Nenhum ticket com status '{filtro_status}'")
        else:
            st.info("📭 Você ainda não abriu nenhum ticket de suporte.")

# 
# 🔄 ROTEAMENTO PRINCIPAL
# 

try:
    st.sidebar.image("https://cdn-icons-png.flaticon.com/512/1055/1055644.png", width=60)
    st.sidebar.title("Menu")
    st.sidebar.write(f"👤 {st.session_state.usuario} ({st.session_state.tipo_usuario.upper()})")

    if st.session_state.cnpj_cpf:
        st.sidebar.write(f"🏢 {st.session_state.cnpj_cpf}")

    if st.session_state.tipo_usuario == 'admin':
        opcoes = ["📦 Entrada", "📦 Saída", "📄 Notas", "⚙️ Admin"]
    elif st.session_state.tipo_usuario == 'gerente':
        opcoes = ["📦 Entrada", "📦 Saída", "📄 Notas", "👥 Gerenciar"]
    else:
        opcoes = ["📦 Entrada", "📦 Saída", "📄 Notas", "🆘 Suporte"]

    pagina = st.sidebar.radio("Navegação:", opcoes)

    st.sidebar.divider()
    if st.sidebar.button("🚪 Logout", use_container_width=True):
        registrar_log(st.session_state.usuario_id, st.session_state.tipo_usuario, 
                     st.session_state.cnpj_cpf if st.session_state.cnpj_cpf else '', 'LOGOUT')
        st.session_state.autenticado = False
        st.session_state.usuario = None
        st.session_state.tipo_usuario = None
        st.session_state.cnpj_cpf = None
        st.session_state.usuario_id = None
        

    st.sidebar.caption("🔒 Protegido")

    if pagina == "📦 Entrada":
        app_auditoria_itens_entrada()
    elif pagina == "📦 Saída":
        app_auditoria_itens_saida()
    elif pagina == "📄 Notas":
        app_auditoria_notas()
    elif pagina == "⚙️ Admin":
        app_painel_admin()
    elif pagina == "👥 Gerenciar":
        app_painel_gerente()
    elif pagina == "🆘 Suporte":
        app_painel_suporte_usuario()
except st.runtime.scriptrunner.StopException:
    raise
except Exception as _route_err:
    import traceback as _tb
    _route_tb = _tb.format_exc()
    print(f"❌ [RUNTIME] Erro no roteamento principal: {str(_route_err)}")
    print(_route_tb)
    st.error(f"❌ Erro inesperado na aplicação: {str(_route_err)}")
    st.code(_route_tb, language="python")